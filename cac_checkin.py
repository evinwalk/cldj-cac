#!/usr/bin/env python3
"""
DoD CAC Check-In / Check-Out System  —  v2
Standalone Python, zero external dependencies.

Features:
  - FASC-N (18-char alphanumeric) parsing per NIST SP 800-73-4
  - SQLite roster DB  (fascn -> personnel details)
  - CSV roster import
  - Unknown-card modal -> manual entry -> cached for future scans
  - Auto IN/OUT direction detection
  - Log export to CSV and XLSX (pure stdlib)
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import csv
import re
import os
import datetime
import pathlib
import zipfile
import calendar

# Leave type constants
LEAVE_RR = "R&R LEAVE"
LEAVE_96 = "96-HOUR LIBERTY"
# === LEAVE FEATURE PATCH APPLIED v6 ===

DB_PATH  = pathlib.Path("cac_roster.db")

# =============================================================================
# FASC-N PARSER  (NIST SP 800-73-4 Appendix A)
# =============================================================================

class FASCNParser:
    """
    The FASC-N stored on a CAC is a 200-bit (25-byte) BCD value.
    Card readers typically emit it as a 25-digit decimal string, an
    18-character alphanumeric encoding, or sometimes hex.

    Field layout (decimal digits):
      agency_code          4 digits
      system_code          4 digits
      credential_number    6 digits
      cs                   1 digit   (credential series)
      ici                  1 digit   (individual credential issue)
      pi                  10 digits  (person identifier)
      oc                   1 digit   (org category)
      oi                   4 digits  (org identifier)
      poa                  1 digit   (person/org assoc)

    Front PDF417 barcode:
      The front of a CAC has a PDF417 barcode that embeds personnel data
      (name, branch, rank) along with the FASC-N in a proprietary variable-
      length encoding. The raw output is 30-80 chars, mixed case, unstable
      between scans. We detect it by length + pattern, then extract the
      stable 18-char FASC-N tail to use as the roster lookup key.
      The FASC-N always appears as the last 7-10 uppercase alphanum chars
      of a longer segment that begins with a known prefix pattern.
    """

    # Known stable prefix shared across front PDF417 scans of DoD CACs
    # (reader-specific; derived from observed samples)
    FRONT_PDF417_MIN_LEN = 25   # shorter than this = back barcode or manual
    FRONT_PDF417_MAX_LEN = 120  # longer than this = something else entirely

    @staticmethod
    def normalise(raw: str) -> str:
        return re.sub(r"[\s\-]", "", raw).upper()

    @classmethod
    def is_front_pdf417(cls, raw: str) -> bool:
        """
        Detect front PDF417 output. Characteristics observed across scans:
          - Length 25-120 chars (after stripping spaces)
          - Mixed alphanum — contains letters AND digits
          - NOT a clean 18-char back barcode
          - Contains embedded name-like alpha sequences (3+ consecutive letters)
        """
        stripped = re.sub(r"\s", "", raw)
        n = stripped.upper()
        if not (cls.FRONT_PDF417_MIN_LEN <= len(n) <= cls.FRONT_PDF417_MAX_LEN):
            return False
        if len(n) == 18:          # exact 18 = clean back barcode, not front
            return False
        has_letters = bool(re.search(r"[A-Z]{3,}", n))
        has_digits  = bool(re.search(r"\d", n))
        return has_letters and has_digits

    @classmethod
    def extract_fascn_from_front(cls, raw: str) -> str | None:
        """
        Extract the stable 18-char FASC-N from a front PDF417 blob.

        Strategy: the FASC-N encoded on the back is also embedded in the
        front barcode blob. Across both observed scans, the tail of the
        blob ends with an uppercase alphanum segment that overlaps with or
        contains the back-barcode FASC-N. We look for the last run of 14-20
        uppercase alphanumeric chars at the end of the stripped string and
        return the rightmost 18 chars as the candidate FASC-N key.

        If the blob is shorter, fall back to the full stripped string.
        """
        n = re.sub(r"\s", "", raw).upper()
        # Find the last contiguous alphanum segment of 14+ chars
        segments = re.findall(r"[A-Z0-9]{14,}", n)
        if segments:
            tail = segments[-1]
            # Prefer exactly 18 chars; if longer, take rightmost 18
            if len(tail) >= 18:
                return tail[-18:]
            return tail
        # Fallback: strip spaces, return full normalised string
        return n if len(n) >= 10 else None

    @classmethod
    def parse(cls, raw: str) -> dict:
        # ── Front PDF417 detection ──────────────────────────────────────
        if cls.is_front_pdf417(raw):
            fascn_key = cls.extract_fascn_from_front(raw)
            if fascn_key:
                return {
                    "raw":      re.sub(r"\s", "", raw).upper(),
                    "source":   "front_PDF417",
                    "fascn_key": fascn_key,
                    "front_blob": raw.strip(),
                }
            # Could not extract — fall through to generic handling

        n = cls.normalise(raw)
        result = {"raw": n, "source": "FASC-N"}

        # Hex (50 chars) -> convert to decimal string
        if re.fullmatch(r"[0-9A-F]{50}", n):
            try:
                dec = str(int(n, 16))
                n = dec.zfill(32)
            except ValueError:
                pass

        digits = re.sub(r"\D", "", n)

        if len(digits) >= 25:
            d = digits[:32] if len(digits) >= 32 else digits.zfill(32)
            result.update({
                "agency_code":       d[0:4],
                "system_code":       d[4:8],
                "credential_number": d[8:14],
                "credential_series": d[14:15],
                "ici":               d[15:16],
                "person_identifier": d[16:26],
                "org_category":      d[26:27],
                "org_identifier":    d[27:31],
                "poa":               d[31:32] if len(d) > 31 else "",
            })
        elif len(n) == 18:
            result["credential_number"] = n[4:10] if len(n) >= 10 else n
            result["agency_code"]       = n[0:4]
        else:
            result["credential_number"] = n

        result["fascn_key"] = n
        return result

    @staticmethod
    def is_fascn(raw: str) -> bool:
        n = re.sub(r"[\s\-]", "", raw)
        return (bool(re.fullmatch(r"[0-9A-Za-z]{16,50}", n)) and not n.isdigit()) or \
               (n.isdigit() and 18 <= len(n) <= 32)


import base64

def _normalise_dob(raw: str) -> str:
    """
    Accept common date formats and normalise to YYYY-MM-DD.
      MM/DD/YYYY  →  1993-10-19
      MMDDYYYY    →  1993-10-19
      YYYY-MM-DD  →  (passthrough)
    """
    raw = raw.strip()
    # MM/DD/YYYY
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", raw)
    if m:
        return f"{m.group(3)}-{m.group(1).zfill(2)}-{m.group(2).zfill(2)}"
    # MMDDYYYY
    m = re.fullmatch(r"(\d{2})(\d{2})(\d{4})", raw)
    if m:
        return f"{m.group(3)}-{m.group(1)}-{m.group(2)}"
    # Already ISO
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
        return raw
    return raw  # passthrough unknown formats


# =============================================================================
# DATABASE  (SQLite, stdlib only)
# =============================================================================

class RosterDB:
    SCHEMA = """
    CREATE TABLE IF NOT EXISTS personnel (
        fascn_key        TEXT PRIMARY KEY,
        raw_fascn        TEXT,
        edipi            TEXT UNIQUE,
        last_name        TEXT,
        first_name       TEXT,
        middle_name      TEXT,
        dob              TEXT,
        place_of_birth   TEXT,
        gender           TEXT,
        branch           TEXT,
        rank             TEXT,
        affiliation      TEXT,
        added_by         TEXT DEFAULT 'manual',
        added_at         TEXT,
        liberty_risk     INTEGER DEFAULT 0,
        liberty_risk_reason TEXT DEFAULT ''
    );

    CREATE TABLE IF NOT EXISTS access_log (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp        TEXT,
        date             TEXT,
        time             TEXT,
        direction        TEXT,
        fascn_key        TEXT,
        edipi            TEXT,
        last_name        TEXT,
        first_name       TEXT,
        middle_name      TEXT,
        branch           TEXT,
        rank             TEXT,
        affiliation      TEXT,
        destination      TEXT
    );

    CREATE TABLE IF NOT EXISTS roster_audit_log (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp        TEXT,
        action           TEXT,
        fascn_key        TEXT,
        edipi            TEXT,
        last_name        TEXT,
        first_name       TEXT,
        changed_fields   TEXT,
        old_values       TEXT,
        new_values       TEXT,
        performed_by     TEXT DEFAULT 'operator'
    );

    CREATE TABLE IF NOT EXISTS liberty_risk_log (
        id                    INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp             TEXT,
        fascn_key             TEXT,
        edipi                 TEXT,
        last_name             TEXT,
        first_name            TEXT,
        direction             TEXT,
        destination           TEXT,
        acknowledger_name     TEXT,
        acknowledger_edipi    TEXT,
        liberty_risk_reason   TEXT
    );
    
    CREATE TABLE IF NOT EXISTS leave_records (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        fascn_key     TEXT,
        edipi         TEXT,
        last_name     TEXT,
        first_name    TEXT,
        leave_type    TEXT,
        start_date    TEXT,
        end_date      TEXT,
        status        TEXT DEFAULT 'ACTIVE',
        departure_ts  TEXT,
        return_ts     TEXT
    );
    """

    # Column aliases: maps every known CSV/XLSX header variant → internal field name
    # Covers original source CSV, JPERSTAT XLSX/CSV, and generic formats.
    COL_MAP = {
        # ── Original source CSV ───────────────────────────────────────────────
        "subject_last_name":  "last_name",
        "subject_first_name": "first_name",
        "date_of_birth":      "dob",
        "place_of_birth":     "place_of_birth",
        "place_of_birth_":    "place_of_birth",
        "edipi":              "edipi",
        # ── JPERSTAT XLSX/CSV columns (normalised) ────────────────────────────
        "leave_blank":        "_skip",
        "unit":               "unit",
        "blood_type":         "blood_type",
        "ln":                 "last_name",
        "fn":                 "first_name",
        "compo":              "affiliation",
        "svc":                "branch",
        "grade":              "rank",
        "sex":                "gender",
        "pdy":                "pdy_status",
        "current_country/3_letter_country_code": "current_country",
        "current_country":    "current_country",
        "current_city/state/_country":           "current_city",
        "current_city/state/country":            "current_city",
        "current_city":       "current_city",
        "posture_location":   "posture_location",
        "tad/leave_start":    "tad_start",
        "tad/leave_end":      "tad_end",
        "theater_arrival":    "theater_arrival",
        "deros":              "deros",
        "citizenship_(us,_dj_nat'l,_foreign_nat'l)": "citizenship",
        "citizenship":        "citizenship",
        "remarks":            "remarks",
        "live_on_camp_(y/n)": "live_on_camp",
        "live_on_camp":       "live_on_camp",
        # ── Generic alternates ────────────────────────────────────────────────
        "last_name":          "last_name",
        "first_name":         "first_name",
        "middle_name":        "middle_name",
        "fascn":              "fascn",
        "fascn_key":          "fascn",
        "fasc_n":             "fascn",
        "dob":                "dob",
        "gender":             "gender",
        "branch":             "branch",
        "rank":               "rank",
        "affiliation":        "affiliation",
    }

    # JPERSTAT signature: at least 3 of these must appear to identify the format.
    # 'EDIPI' alone is not enough since legacy CSVs also have it.
    _JPERSTAT_MARKERS = {"LN", "FN", "GRADE", "SVC", "DEROS", "COMPO", "PDY"}
    _JPERSTAT_MIN_MATCH = 3

    def __init__(self, path: pathlib.Path = DB_PATH):
        self.path = path
        self.conn = sqlite3.connect(str(path), check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.conn.executescript(self.SCHEMA)
        self._migrate()
        self.conn.commit()

    def _migrate(self):
        """Add columns introduced after initial schema without wiping data."""
        cur = self.conn.execute("PRAGMA table_info(personnel)")
        existing = {row[1] for row in cur.fetchall()}
        additions = {
            "place_of_birth":      "TEXT DEFAULT ''",
            "destination":         "TEXT DEFAULT ''",
            "liberty_risk":        "INTEGER DEFAULT 0",
            "liberty_risk_reason": "TEXT DEFAULT ''",
            "unit":                "TEXT DEFAULT ''",
        }
        for col, defn in additions.items():
            if col not in existing:
                self.conn.execute(
                    f"ALTER TABLE personnel ADD COLUMN {col} {defn}"
                )
        try:
            self.conn.execute("ALTER TABLE personnel DROP COLUMN ssn_enc")
        except Exception:
            try:
                self.conn.execute("UPDATE personnel SET ssn_enc=''")
            except Exception:
                pass
        try:
            acur = self.conn.execute("PRAGMA table_info(access_log)")
            acols = {r[1] for r in acur.fetchall()}
            if "group_id" not in acols:
                self.conn.execute(
                    "ALTER TABLE access_log ADD COLUMN group_id TEXT DEFAULT ''")
        except Exception:
            pass
        # Ensure audit and liberty risk tables exist on older DBs
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS roster_audit_log (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp        TEXT,
                action           TEXT,
                fascn_key        TEXT,
                edipi            TEXT,
                last_name        TEXT,
                first_name       TEXT,
                changed_fields   TEXT,
                old_values       TEXT,
                new_values       TEXT,
                performed_by     TEXT DEFAULT 'operator'
            );
            CREATE TABLE IF NOT EXISTS liberty_risk_log (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp             TEXT,
                fascn_key             TEXT,
                edipi                 TEXT,
                last_name             TEXT,
                first_name            TEXT,
                direction             TEXT,
                destination           TEXT,
                acknowledger_name     TEXT,
                acknowledger_edipi    TEXT,
                liberty_risk_reason   TEXT
            );
            CREATE TABLE IF NOT EXISTS leave_records (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                fascn_key     TEXT,
                edipi         TEXT,
                last_name     TEXT,
                first_name    TEXT,
                leave_type    TEXT,
                start_date    TEXT,
                end_date      TEXT,
                status        TEXT DEFAULT 'ACTIVE',
                departure_ts  TEXT,
                return_ts     TEXT
            );
        """)

    def _audit(self, action: str, fascn_key: str, data: dict,
               old: dict = None, changed_fields: list = None,
               performed_by: str = "operator"):
        """Write one row to roster_audit_log."""
        now = datetime.datetime.now().isoformat(timespec="seconds")
        old_vals = ""
        new_vals = ""
        if old and changed_fields:
            old_vals = "; ".join(f"{f}={old.get(f,'')}" for f in changed_fields)
            new_vals = "; ".join(f"{f}={data.get(f,'')}" for f in changed_fields)
        self.conn.execute("""
            INSERT INTO roster_audit_log
              (timestamp, action, fascn_key, edipi, last_name, first_name,
               changed_fields, old_values, new_values, performed_by)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            now, action, fascn_key,
            data.get("edipi", ""),
            data.get("last_name", ""),
            data.get("first_name", ""),
            ", ".join(changed_fields) if changed_fields else "",
            old_vals, new_vals, performed_by,
        ))
        self.conn.commit()


    def lookup(self, fascn_key: str):
        """
        Primary lookup by exact fascn_key.
        Falls back to substring matching — front and back barcodes encode
        the same FASC-N differently but share a stable 6+ char common
        substring (derived from the credential number embedded in both).
        We extract candidate substrings from the query key and search stored
        keys, then vice versa, using LIKE for efficiency.
        """
        # 1. Exact match
        cur = self.conn.execute(
            "SELECT * FROM personnel WHERE fascn_key = ?", (fascn_key,)
        )
        row = cur.fetchone()
        if row:
            return dict(row)

        # 2. Check if any stored key is a substring of the query key
        #    (back key embedded in longer front blob key)
        cur = self.conn.execute("SELECT * FROM personnel")
        for row in cur.fetchall():
            stored = row["fascn_key"]
            # stored key appears inside the query key
            if len(stored) >= 10 and stored in fascn_key:
                return dict(row)
            # query key appears inside the stored key
            if len(fascn_key) >= 10 and fascn_key in stored:
                return dict(row)

        # 3. Sliding-window shared substring match (min 6 chars)
        #    Extracts all 6-char windows from the query key and checks
        #    if any stored key contains that window — catches partial
        #    front/back encoding overlaps like 'TP6Q6B'
        if len(fascn_key) >= 6:
            cur = self.conn.execute("SELECT * FROM personnel")
            stored_rows = [dict(r) for r in cur.fetchall()]
            for window_len in range(min(len(fascn_key), 10), 5, -1):
                for start in range(len(fascn_key) - window_len + 1):
                    window = fascn_key[start:start + window_len]
                    for row in stored_rows:
                        if window in row["fascn_key"]:
                            return row

        return None

    def lookup_exact(self, fascn_key: str):
        """Exact-match only — used internally."""
        cur = self.conn.execute(
            "SELECT * FROM personnel WHERE fascn_key = ?", (fascn_key,)
        )
        row = cur.fetchone()
        return dict(row) if row else None

    def upsert_person(self, fascn_key: str, data: dict,
                      added_by: str = "manual", performed_by: str = "operator"):
        now = datetime.datetime.now().isoformat(timespec="seconds")
        # Check if record already exists to determine INSERT vs UPDATE
        existing = self.lookup_exact(fascn_key)
        self.conn.execute("""
            INSERT INTO personnel
              (fascn_key, raw_fascn, edipi, last_name, first_name, middle_name,
               dob, place_of_birth, gender, branch, rank, affiliation,
               added_by, added_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(fascn_key) DO UPDATE SET
              edipi=excluded.edipi,
              last_name=excluded.last_name,
              first_name=excluded.first_name,
              middle_name=excluded.middle_name,
              dob=excluded.dob,
              place_of_birth=excluded.place_of_birth,
              gender=excluded.gender,
              branch=excluded.branch,
              rank=excluded.rank,
              affiliation=excluded.affiliation,
              added_by=excluded.added_by,
              added_at=excluded.added_at
        """, (
            fascn_key,
            data.get("raw_fascn", fascn_key),
            data.get("edipi", ""),
            data.get("last_name", ""),
            data.get("first_name", ""),
            data.get("middle_name", ""),
            data.get("dob", ""),
            data.get("place_of_birth", ""),
            data.get("gender", ""),
            data.get("branch", ""),
            data.get("rank", ""),
            data.get("affiliation", ""),
            added_by,
            now,
        ))
        self.conn.commit()
        # Audit
        if existing:
            editable = ["last_name","first_name","middle_name","edipi",
                        "dob","place_of_birth","gender","branch","rank","affiliation"]
            changed = [f for f in editable if str(existing.get(f,"")) != str(data.get(f,""))]
            if changed:
                self._audit("UPDATE", fascn_key, data, old=existing,
                            changed_fields=changed, performed_by=performed_by)
        else:
            self._audit("INSERT", fascn_key, data, performed_by=performed_by)

    def update_person(self, fascn_key: str, data: dict, performed_by: str = "operator"):
        """Update editable fields on an existing record."""
        existing = self.lookup_exact(fascn_key)
        if not existing:
            return
        editable = ["last_name","first_name","middle_name","edipi",
                    "dob","place_of_birth","gender","branch","rank","affiliation"]
        changed = [f for f in editable if str(existing.get(f,"")) != str(data.get(f,""))]
        if not changed:
            return
        self.conn.execute("""
            UPDATE personnel SET
              last_name=?, first_name=?, middle_name=?, edipi=?,
              dob=?, place_of_birth=?, gender=?, branch=?, rank=?,
              affiliation=?
            WHERE fascn_key=?
        """, (
            data.get("last_name",   existing["last_name"]),
            data.get("first_name",  existing["first_name"]),
            data.get("middle_name", existing.get("middle_name","")),
            data.get("edipi",       existing.get("edipi","")),
            data.get("dob",         existing.get("dob","")),
            data.get("place_of_birth", existing.get("place_of_birth","")),
            data.get("gender",      existing.get("gender","")),
            data.get("branch",      existing.get("branch","")),
            data.get("rank",        existing.get("rank","")),
            data.get("affiliation", existing.get("affiliation","")),
            fascn_key,
        ))
        self.conn.commit()
        self._audit("UPDATE", fascn_key, data, old=existing,
                    changed_fields=changed, performed_by=performed_by)

    def delete_person(self, fascn_key: str, performed_by: str = "operator"):
        """Soft-delete: remove from personnel but keep in audit log."""
        existing = self.lookup_exact(fascn_key)
        if not existing:
            return
        self._audit("DELETE", fascn_key, existing, performed_by=performed_by)
        self.conn.execute("DELETE FROM personnel WHERE fascn_key=?", (fascn_key,))
        self.conn.commit()

    def get_audit_log(self, text_filter: str = "") -> list:
        q = "SELECT * FROM roster_audit_log"
        params = []
        if text_filter:
            q += (" WHERE last_name LIKE ? OR first_name LIKE ? "
                  "OR fascn_key LIKE ? OR edipi LIKE ? OR action LIKE ?")
            t = f"%{text_filter}%"
            params = [t, t, t, t, t]
        q += " ORDER BY id DESC"
        return [dict(r) for r in self.conn.execute(q, params).fetchall()]

    def audit_count(self) -> int:
        return self.conn.execute(
            "SELECT COUNT(*) FROM roster_audit_log").fetchone()[0]

    def set_liberty_risk(self, fascn_key: str, is_risk: bool,
                         reason: str = "", performed_by: str = "operator"):
        """Flag or unflag a person as a Liberty Risk."""
        existing = self.lookup_exact(fascn_key)
        if not existing:
            return
        self.conn.execute(
            "UPDATE personnel SET liberty_risk=?, liberty_risk_reason=? "
            "WHERE fascn_key=?",
            (1 if is_risk else 0, reason.strip(), fascn_key)
        )
        self.conn.commit()
        action = "LIBERTY_RISK_SET" if is_risk else "LIBERTY_RISK_CLEARED"
        self._audit(action, fascn_key,
                    {**dict(existing), "liberty_risk_reason": reason},
                    old=dict(existing),
                    changed_fields=["liberty_risk", "liberty_risk_reason"],
                    performed_by=performed_by)

    def log_liberty_risk_acknowledgement(self, fascn_key: str, person: dict,
                                          direction: str, destination: str,
                                          acknowledger_name: str,
                                          acknowledger_edipi: str):
        """Record that an operator acknowledged a Liberty Risk scan event."""
        now = datetime.datetime.now().isoformat(timespec="seconds")
        self.conn.execute("""
            INSERT INTO liberty_risk_log
              (timestamp, fascn_key, edipi, last_name, first_name,
               direction, destination, acknowledger_name,
               acknowledger_edipi, liberty_risk_reason)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            now,
            fascn_key,
            person.get("edipi", ""),
            person.get("last_name", ""),
            person.get("first_name", ""),
            direction,
            destination,
            acknowledger_name.strip(),
            acknowledger_edipi.strip(),
            person.get("liberty_risk_reason", ""),
        ))
        self.conn.commit()

    def get_liberty_risk_log(self, text_filter: str = "") -> list:
        q = "SELECT * FROM liberty_risk_log"
        params = []
        if text_filter:
            q += (" WHERE last_name LIKE ? OR first_name LIKE ? "
                  "OR acknowledger_name LIKE ? OR acknowledger_edipi LIKE ?")
            t = f"%{text_filter}%"
            params = [t, t, t, t]
        q += " ORDER BY id DESC"
        return [dict(r) for r in self.conn.execute(q, params).fetchall()]

    def liberty_risk_log_count(self) -> int:
        return self.conn.execute(
            "SELECT COUNT(*) FROM liberty_risk_log").fetchone()[0]

    # ── Leave records ─────────────────────────────────────────────────────────

    def create_leave_record(self, fascn_key: str, person: dict,
                            leave_type: str, start_date: str, end_date: str):
        """Create an ACTIVE leave record and stamp the departure time."""
        now = datetime.datetime.now().isoformat(timespec="seconds")
        self.conn.execute("""
            INSERT INTO leave_records
              (fascn_key, edipi, last_name, first_name, leave_type,
               start_date, end_date, status, departure_ts)
            VALUES (?,?,?,?,?,?,?, 'ACTIVE', ?)
        """, (
            fascn_key,
            person.get("edipi", ""),
            person.get("last_name", ""),
            person.get("first_name", ""),
            leave_type,
            start_date,
            end_date,
            now,
        ))
        self.conn.commit()

    def get_active_leave(self, fascn_key: str):
        """Return the most recent ACTIVE leave record for this person, or None."""
        cur = self.conn.execute(
            "SELECT * FROM leave_records WHERE fascn_key=? AND status='ACTIVE' "
            "ORDER BY id DESC LIMIT 1", (fascn_key,)
        )
        row = cur.fetchone()
        return dict(row) if row else None

    def mark_leave_returned(self, leave_id: int):
        """Mark a leave record RETURNED and stamp the return time."""
        now = datetime.datetime.now().isoformat(timespec="seconds")
        self.conn.execute(
            "UPDATE leave_records SET status='RETURNED', return_ts=? WHERE id=?",
            (now, leave_id)
        )
        self.conn.commit()

    def get_leave_records(self, status_filter: str = "ALL",
                          text_filter: str = "") -> list:
        q = "SELECT * FROM leave_records"
        params = []
        clauses = []
        if status_filter != "ALL":
            clauses.append("status=?")
            params.append(status_filter)
        if text_filter:
            clauses.append("(last_name LIKE ? OR first_name LIKE ? "
                           "OR edipi LIKE ? OR leave_type LIKE ?)")
            t = f"%{text_filter}%"
            params += [t, t, t, t]
        if clauses:
            q += " WHERE " + " AND ".join(clauses)
        q += " ORDER BY id DESC"
        return [dict(r) for r in self.conn.execute(q, params).fetchall()]

    def leave_active_count(self) -> int:
        return self.conn.execute(
            "SELECT COUNT(*) FROM leave_records WHERE status='ACTIVE'"
        ).fetchone()[0]

    def leave_total_count(self) -> int:
        return self.conn.execute(
            "SELECT COUNT(*) FROM leave_records").fetchone()[0]

    def get_leave_record(self, leave_id: int):
        cur = self.conn.execute(
            "SELECT * FROM leave_records WHERE id=?", (leave_id,))
        row = cur.fetchone()
        return dict(row) if row else None

    def update_leave_dates(self, leave_id: int, start_date: str, end_date: str):
        self.conn.execute(
            "UPDATE leave_records SET start_date=?, end_date=? WHERE id=?",
            (start_date, end_date, leave_id))
        self.conn.commit()

    def reopen_leave(self, leave_id: int):
        # Undo a return: set status back to ACTIVE and clear the return time.
        self.conn.execute(
            "UPDATE leave_records SET status='ACTIVE', return_ts=NULL WHERE id=?",
            (leave_id,))
        self.conn.commit()

    def delete_leave_record(self, leave_id: int):
        self.conn.execute("DELETE FROM leave_records WHERE id=?", (leave_id,))
        self.conn.commit()

    def update_unit(self, fascn_key: str, unit: str):
        self.conn.execute("UPDATE personnel SET unit=? WHERE fascn_key=?",
                          (unit, fascn_key))
        self.conn.commit()




    # ── Import helpers ────────────────────────────────────────────────────────

    @staticmethod
    def _excel_serial_to_date(val) -> str:
        """
        Convert an Excel serial date (integer or float) or datetime object
        to ISO YYYY-MM-DD.  Excel epoch = 1899-12-30 (with the Lotus 1-2-3
        leap-year bug baked in — day 60 doesn't exist, so offset by 1 for
        serials > 59).
        """
        if val is None:
            return ""
        if isinstance(val, datetime.datetime):
            return val.strftime("%Y-%m-%d")
        if isinstance(val, datetime.date):
            return val.strftime("%Y-%m-%d")
        try:
            serial = int(float(str(val)))
            if serial <= 0:
                return ""
            # Excel epoch
            epoch = datetime.date(1899, 12, 30)
            d = epoch + datetime.timedelta(days=serial)
            return d.strftime("%Y-%m-%d")
        except (ValueError, TypeError, OverflowError):
            return _normalise_dob(str(val))

    @staticmethod
    def _norm_col(raw: str) -> str:
        """Normalise a column header for COL_MAP lookup."""
        return (raw.replace("\xa0", " ")
                   .replace("\n", " ")
                   .strip()
                   .lower()
                   .replace(" ", "_")
                   .rstrip("_"))

    def _is_jperstat(self, headers: list) -> bool:
        """Return True if the header row looks like a JPERSTAT sheet (≥3 unique markers)."""
        header_set = {str(h).strip().upper() for h in headers if h}
        return len(self._JPERSTAT_MARKERS & header_set) >= self._JPERSTAT_MIN_MATCH

    def _build_record_from_row(self, r: dict) -> dict:
        """
        Given a normalised row dict (keys already mapped via COL_MAP),
        return a data dict ready for upsert_person.
        Handles JPERSTAT-specific fields (DEROS as ERD, GRADE as rank, etc.)
        """
        last  = r.get("last_name", "").strip().title()
        first = r.get("first_name", "").strip().title()
        edipi = re.sub(r"\D", "", r.get("edipi", ""))
        fascn = r.get("fascn", "").strip()

        # Date priority: DEROS > THEATER ARRIVAL > DOB
        # DEROS is the End of Required Overseas Service date — primary date field
        # THEATER ARRIVAL is used as fallback (date they arrived in theatre)
        deros_raw   = r.get("deros", "")
        arrival_raw = r.get("theater_arrival", "")
        if deros_raw:
            date_field = self._excel_serial_to_date(deros_raw)
        elif arrival_raw:
            date_field = self._excel_serial_to_date(arrival_raw)
        else:
            date_field = r.get("dob", "")

        # Gender: M/F/X
        gender_raw = r.get("gender", "").strip().upper()
        gender = gender_raw if gender_raw in ("M", "F", "X") else ""

        # Branch: JPERSTAT SVC field
        branch = r.get("branch", "").strip().upper()
        if branch in ("N/A", "NA", ""):
            branch = r.get("affiliation", "").strip()

        rank        = r.get("rank", "").strip().upper()
        affiliation = r.get("affiliation", "").strip()

        # Citizenship — also accept keys that didn't fully normalise
        citizenship = r.get("citizenship", "")
        if not citizenship:
            # Search for any key that starts with 'citizenship'
            for k, v in r.items():
                if k.startswith("citizenship") and v:
                    citizenship = v
                    break
        citizenship = citizenship.strip()

        place_of_birth = (r.get("place_of_birth", "") or
                          r.get("current_city", "") or "").strip()
        unit    = r.get("unit", "").strip()
        remarks = r.get("remarks", "").strip()

        return {
            "last_name":      last,
            "first_name":     first,
            "edipi":          edipi,
            "fascn":          fascn,
            "dob":            r.get("dob", ""),
            "deros":          date_field,
            "gender":         gender,
            "branch":         branch,
            "rank":           rank,
            "affiliation":    affiliation,
            "place_of_birth": place_of_birth,
            "citizenship":    citizenship,
            "unit":           unit,
            "remarks":        remarks,
            "middle_name":    r.get("middle_name", ""),
        }

    def import_file(self, path: str):
        """
        Universal import — detects format from extension and content:
          • .xlsx / .xlsm  →  import_jperstat_xlsx (PERSTAT sheet)
          • .csv / .tsv    →  auto-detects JPERSTAT CSV vs legacy CSV

        Returns (imported, skipped, format_detected).
        """
        ext = pathlib.Path(path).suffix.lower()
        if ext in (".xlsx", ".xlsm", ".xls"):
            return self._import_xlsx(path)
        else:
            return self._import_csv_file(path)

    # Keep old name as alias for backward compatibility
    def import_csv(self, path: str):
        imp, sk, _ = self.import_file(path)
        return imp, sk

    def _import_xlsx(self, path: str):
        """Import from JPERSTAT XLSX — reads the PERSTAT sheet."""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError(
                "openpyxl is required to import XLSX files.\n"
                "Run: pip install openpyxl"
            )

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        # Find the right sheet: prefer PERSTAT, then first sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "PERSTAT" in name.upper() or "ROSTER" in name.upper():
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return 0, 0, "xlsx_empty"

        # Header row: find it by looking for EDIPI/LN/FN markers
        header_row_idx = 0
        for i, row in enumerate(rows[:5]):
            if self._is_jperstat(list(row)):
                header_row_idx = i
                break

        headers = [str(h).strip() if h else "" for h in rows[header_row_idx]]
        data_rows = rows[header_row_idx + 1:]
        fmt = "jperstat_xlsx"

        imported = skipped = 0
        for raw_row in data_rows:
            # Build normalised dict from row values
            r: dict = {}
            for i, val in enumerate(raw_row):
                if i >= len(headers):
                    break
                header = headers[i]
                if not header:
                    continue
                norm = self._norm_col(header)
                mapped = self.COL_MAP.get(norm, norm)
                if mapped == "_skip":
                    continue
                # Clean value
                if val is None:
                    r[mapped] = ""
                elif isinstance(val, (datetime.datetime, datetime.date)):
                    r[mapped] = self._excel_serial_to_date(val)
                else:
                    r[mapped] = str(val).replace("\xa0", " ").strip()

            rec = self._build_record_from_row(r)
            imp, sk = self._upsert_record(rec)
            imported += imp
            skipped  += sk

        return imported, skipped, fmt

    def _import_csv_file(self, path: str):
        """Import from CSV — detects JPERSTAT CSV vs legacy CSV by headers."""
        raw_bytes = open(path, "rb").read()
        encoding = "utf-8-sig"
        for enc in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                raw_bytes.decode(enc)
                encoding = enc
                break
            except (UnicodeDecodeError, LookupError):
                continue

        imported = skipped = 0
        fmt = "unknown_csv"

        with open(path, newline="", encoding=encoding, errors="replace") as f:
            sample = f.read(4096)
            f.seek(0)
            delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
            reader = csv.DictReader(f, delimiter=delimiter)

            if reader.fieldnames:
                fmt = ("jperstat_csv"
                       if self._is_jperstat(list(reader.fieldnames))
                       else "legacy_csv")

            for row in reader:
                r: dict = {}
                for raw_col, val in row.items():
                    if raw_col is None:
                        continue
                    norm = self._norm_col(raw_col)
                    mapped = self.COL_MAP.get(norm, norm)
                    if mapped == "_skip":
                        continue
                    r[mapped] = val.replace("\xa0", " ").strip() if val else ""

                rec = self._build_record_from_row(r)
                imp, sk = self._upsert_record(rec)
                imported += imp
                skipped  += sk

        return imported, skipped, fmt

    def _upsert_record(self, rec: dict):
        """
        Resolve the DB key and call upsert_person.
        Returns (1, 0) on success, (0, 1) on skip.
        Key priority:
          1. FASC-N  (if present in the import — rare)
          2. EDIPI   (most common — keyed as 'EDIPI:<edipi>')
          3. Skip    (no usable identifier)
        """
        last  = rec.get("last_name", "")
        first = rec.get("first_name", "")
        edipi = rec.get("edipi", "")
        fascn = rec.get("fascn", "")

        if not last and not first:
            return 0, 1

        if fascn:
            parsed = FASCNParser.parse(fascn)
            key = parsed["fascn_key"]
        elif edipi:
            key = f"EDIPI:{edipi}"
        else:
            return 0, 1

        # Use DEROS as the primary date if no DOB
        dob = rec.get("dob") or rec.get("deros", "")

        self.upsert_person(key, {
            "raw_fascn":      fascn,
            "edipi":          edipi,
            "last_name":      last,
            "first_name":     first,
            "middle_name":    rec.get("middle_name", ""),
            "dob":            dob,
            "place_of_birth": rec.get("place_of_birth", ""),
            "gender":         rec.get("gender", ""),
            "branch":         rec.get("branch", ""),
            "rank":           rec.get("rank", ""),
            "affiliation":    rec.get("affiliation", ""),
        }, added_by="import", performed_by="import")
        self.update_unit(key, rec.get("unit", ""))
        return 1, 0

    def lookup_by_edipi(self, edipi: str):
        """Find a person by EDIPI — used to bridge EDIPI-keyed imports to scanned FASC-Ns."""
        if not edipi:
            return None
        cur = self.conn.execute(
            "SELECT * FROM personnel WHERE edipi = ?", (edipi,)
        )
        row = cur.fetchone()
        return dict(row) if row else None

    def link_fascn_to_edipi(self, fascn_key: str, edipi: str):
        """
        When a card is scanned whose EDIPI matches an EDIPI-keyed import,
        update that record's fascn_key to the real scanned value so future
        scans hit instantly via exact match.
        """
        existing = self.lookup_by_edipi(edipi)
        if not existing:
            return
        old_key = existing["fascn_key"]
        if old_key == fascn_key:
            return  # already linked
        # Update the primary key to the real FASC-N
        try:
            self.conn.execute(
                "UPDATE personnel SET fascn_key=?, raw_fascn=? WHERE fascn_key=?",
                (fascn_key, fascn_key, old_key)
            )
            self.conn.execute(
                "UPDATE access_log SET fascn_key=? WHERE fascn_key=?",
                (fascn_key, old_key)
            )
            self.conn.commit()
            self._audit("FASCN_LINK", fascn_key,
                        {**existing, "edipi": edipi},
                        old={"fascn_key": old_key},
                        changed_fields=["fascn_key"],
                        performed_by="system")
        except sqlite3.IntegrityError:
            pass

    # -- Access log -----------------------------------------------------------

    def roster_count(self) -> int:
        return self.conn.execute("SELECT COUNT(*) FROM personnel").fetchone()[0]

    def log_event(self, direction: str, fascn_key: str, person: dict,
                  destination: str = "", group_id: str = "") -> dict:
        now = datetime.datetime.now()
        entry = {
            "timestamp":   now.isoformat(timespec="seconds"),
            "date":        now.strftime("%Y-%m-%d"),
            "time":        now.strftime("%H:%M:%S"),
            "direction":   direction,
            "fascn_key":   fascn_key,
            "edipi":       person.get("edipi", ""),
            "last_name":   person.get("last_name", ""),
            "first_name":  person.get("first_name", ""),
            "middle_name": person.get("middle_name", ""),
            "branch":      person.get("branch", ""),
            "rank":        person.get("rank", ""),
            "affiliation": person.get("affiliation", ""),
            "destination": destination,
            "group_id":    group_id,
        }
        self.conn.execute("""
            INSERT INTO access_log
              (timestamp, date, time, direction, fascn_key, edipi,
               last_name, first_name, middle_name, branch, rank, affiliation,
               destination, group_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, tuple(entry[k] for k in
              ["timestamp","date","time","direction","fascn_key","edipi",
               "last_name","first_name","middle_name","branch","rank","affiliation",
               "destination","group_id"]))
        self.conn.commit()
        return entry

    def last_group_id(self, fascn_key: str) -> str:
        row = self.conn.execute(
            "SELECT group_id FROM access_log WHERE fascn_key=? AND "
            "direction='OUT' ORDER BY id DESC LIMIT 1", (fascn_key,)).fetchone()
        if not row:
            return ""
        gid = row[0]
        return gid if gid else ""

    def group_members_out(self, group_id: str) -> list:
        if not group_id:
            return []
        keys = [r[0] for r in self.conn.execute(
            "SELECT DISTINCT fascn_key FROM access_log WHERE group_id=?",
            (group_id,)).fetchall()]
        out = []
        for fk in keys:
            if self.last_direction(fk) == "OUT":
                p = self.lookup_exact(fk) or {"fascn_key": fk}
                p = dict(p)
                p["fascn_key"] = fk
                out.append(p)
        return out

    def last_direction(self, fascn_key: str):
        cur = self.conn.execute(
            "SELECT direction FROM access_log WHERE fascn_key=? ORDER BY id DESC LIMIT 1",
            (fascn_key,)
        )
        row = cur.fetchone()
        return row[0] if row else None

    def get_log(self, direction_filter: str = "ALL", text_filter: str = ""):
        q = "SELECT * FROM access_log"
        params = []
        clauses = []
        if direction_filter != "ALL":
            clauses.append("direction=?")
            params.append(direction_filter)
        if text_filter:
            clauses.append(
                "(last_name LIKE ? OR first_name LIKE ? OR "
                "fascn_key LIKE ? OR edipi LIKE ? OR branch LIKE ? "
                "OR destination LIKE ?)"
            )
            t = f"%{text_filter}%"
            params += [t, t, t, t, t, t]
        if clauses:
            q += " WHERE " + " AND ".join(clauses)
        q += " ORDER BY id DESC"
        return [dict(r) for r in self.conn.execute(q, params).fetchall()]

    def clear_log(self):
        self.conn.execute("DELETE FROM access_log")
        self.conn.commit()

    def total_log_count(self) -> int:
        return self.conn.execute("SELECT COUNT(*) FROM access_log").fetchone()[0]

    # -- Exports --------------------------------------------------------------

    EXPORT_COLS = ["timestamp","date","time","direction","fascn_key","edipi",
                   "last_name","first_name","middle_name","branch","rank",
                   "affiliation","destination"]
    EXPORT_HDRS = ["Timestamp","Date","Time","Direction","FASC-N","DoD ID (EDIPI)",
                   "Last Name","First Name","Middle Name","Branch","Rank",
                   "Affiliation","Destination"]

    def export_csv(self, path: str):
        rows = self.get_log()
        with open(path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=self.EXPORT_COLS, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)

    def export_xlsx(self, path: str):
        rows = self.get_log()
        cols, headers = self.EXPORT_COLS, self.EXPORT_HDRS

        def esc(v):
            return (str(v)
                    .replace("&","&amp;").replace("<","&lt;")
                    .replace(">","&gt;").replace('"',"&quot;"))

        def col_letter(n):
            s, n = "", n + 1
            while n:
                n, r = divmod(n - 1, 26)
                s = chr(65 + r) + s
            return s

        all_str, str_idx = [], {}

        def si(val):
            v = str(val)
            if v not in str_idx:
                str_idx[v] = len(all_str)
                all_str.append(v)
            return str_idx[v]

        for h in headers:
            si(h)
        for row in rows:
            for c in cols:
                si(row.get(c, ""))

        rows_xml = []
        cells = "".join(
            f'<c r="{col_letter(ci)}1" t="s"><v>{si(h)}</v></c>'
            for ci, h in enumerate(headers)
        )
        rows_xml.append(f'<row r="1">{cells}</row>')
        for ri, row in enumerate(rows, start=2):
            cells = "".join(
                f'<c r="{col_letter(ci)}{ri}" t="s"><v>{si(row.get(c,""))}</v></c>'
                for ci, c in enumerate(cols)
            )
            rows_xml.append(f'<row r="{ri}">{cells}</row>')

        sst = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            f' count="{len(all_str)}" uniqueCount="{len(all_str)}">'
            + "".join(f'<si><t xml:space="preserve">{esc(s)}</t></si>' for s in all_str)
            + '</sst>'
        )
        sheet = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheetData>' + "".join(rows_xml) + '</sheetData></worksheet>'
        )
        wb = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Access Log" sheetId="1" r:id="rId1"/></sheets></workbook>'
        )
        wb_rels = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            'Target="worksheets/sheet1.xml"/>'
            '<Relationship Id="rId2" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
            'Target="sharedStrings.xml"/>'
            '</Relationships>'
        )
        ct = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '<Override PartName="/xl/sharedStrings.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
            '</Types>'
        )
        rels_root = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
            'Target="xl/workbook.xml"/>'
            '</Relationships>'
        )
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("[Content_Types].xml", ct)
            zf.writestr("_rels/.rels", rels_root)
            zf.writestr("xl/workbook.xml", wb)
            zf.writestr("xl/_rels/workbook.xml.rels", wb_rels)
            zf.writestr("xl/sharedStrings.xml", sst)
            zf.writestr("xl/worksheets/sheet1.xml", sheet)


# =============================================================================
# COLOR / FONT CONSTANTS
# =============================================================================

C = {
    "bg":          "#0A0E1A",
    "surface":     "#111827",
    "surface2":    "#1C2536",
    "border":      "#2A3550",
    "accent":      "#0057B8",
    "accent_light":"#1A6FD4",
    "green":       "#00C27C",
    "red":         "#E03A3A",
    "gold":        "#D4A017",
    "text":        "#E8EDF5",
    "dim":         "#7A8BA8",
    "bright":      "#FFFFFF",
    "in_bg":       "#003D1F",
    "out_bg":      "#3D0000",
    "in_fg":       "#7FFFC0",
    "out_fg":      "#FF9090",
}

FTITLE  = ("Arial", 17, "bold")
FSML    = ("Arial", 9)
FBOLD   = ("Arial", 10, "bold")
FMONO   = ("Courier New", 11)
FMONO10 = ("Courier New", 10)
FID     = ("Courier New", 13, "bold")


# =============================================================================
# UNKNOWN CARD MODAL
# =============================================================================

class UnknownCardModal(tk.Toplevel):
    """
    Shown when a scanned FASC-N is not found in the roster.

    Two modes:
      LOOKUP  — operator enters an EDIPI; if found in DB the record is
                shown for confirmation and the FASC-N is linked to it.
      MANUAL  — operator fills in all fields from scratch to create a
                new record.

    The modal starts in LOOKUP mode.  A "Not in database" button switches
    to MANUAL mode.
    """

    BRANCHES = ["","ARMY","NAVY","USMC","USAF","USSF","USCG",
                "CIVILIAN","CONTRACTOR","OTHER"]
    GENDERS  = ["","M","F","X"]

    def __init__(self, parent, fascn_key: str, parsed: dict,
                 callback, db):
        super().__init__(parent)
        self.title("Unknown Card")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self.callback  = callback
        self.fascn_key = fascn_key
        self.parsed    = parsed
        self.db        = db
        self._matched  = None   # the DB record found via EDIPI lookup

        self._build_header(fascn_key, parsed)
        self._build_lookup_section()
        self._build_manual_section()
        self._build_direction_row()
        self._build_buttons()

        # Start in lookup mode
        self._show_lookup_mode()
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.after(50, self._center)

    # ── Layout helpers ────────────────────────────────────────────────────────

    def _center(self):
        self.update_idletasks()
        pw  = self.master.winfo_rootx()
        py  = self.master.winfo_rooty()
        w   = self.winfo_width()
        h   = self.winfo_height()
        self.geometry(
            f"+{pw + self.master.winfo_width()//2 - w//2}"
            f"+{py + self.master.winfo_height()//2 - h//2}"
        )

    def _build_header(self, fascn_key, parsed):
        # Accent bar
        tk.Frame(self, bg=C["gold"], height=3).pack(fill="x")

        hdr = tk.Frame(self, bg=C["accent"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Unknown Card",
                 font=("Arial", 13, "bold"), fg="white", bg=C["accent"]
                 ).pack(side="left", padx=16)
        self._mode_lbl = tk.Label(hdr, text="",
                 font=("Arial", 9), fg="#BDD7FF", bg=C["accent"])
        self._mode_lbl.pack(side="right", padx=16)

        # FASC-N info strip
        info = tk.Frame(self, bg=C["surface2"], pady=6)
        info.pack(fill="x", padx=14, pady=(10, 0))
        tk.Label(info, text="FASC-N:", font=FSML, fg=C["dim"],
                 bg=C["surface2"]).pack(side="left", padx=8)
        tk.Label(info, text=fascn_key, font=FMONO, fg=C["gold"],
                 bg=C["surface2"]).pack(side="left")
        if parsed.get("agency_code"):
            tk.Label(info, text=f"  Agency: {parsed['agency_code']}",
                     font=FSML, fg=C["dim"], bg=C["surface2"]
                     ).pack(side="left", padx=8)

    def _build_lookup_section(self):
        """EDIPI lookup panel — shown in lookup mode."""
        self._lookup_frame = tk.Frame(self, bg=C["bg"], padx=16, pady=14)

        tk.Label(self._lookup_frame,
                 text="This FASC-N is not in the roster.\n"
                      "If this person is already in the database enter their DoD ID (EDIPI)\n"
                      "below to find their record and link this card to it.",
                 font=("Arial", 9), fg=C["dim"], bg=C["bg"],
                 justify="left").pack(anchor="w", pady=(0, 10))

        edipi_row = tk.Frame(self._lookup_frame, bg=C["bg"])
        edipi_row.pack(fill="x")
        tk.Label(edipi_row, text="DoD ID (EDIPI):", font=FBOLD,
                 fg=C["text"], bg=C["bg"], width=16, anchor="w"
                 ).pack(side="left")
        self._edipi_var = tk.StringVar()
        self._edipi_entry = tk.Entry(
            edipi_row, textvariable=self._edipi_var,
            font=FID, bg=C["surface2"], fg=C["bright"],
            insertbackground=C["bright"], relief="flat",
            highlightbackground=C["border"], highlightthickness=1,
            width=16
        )
        self._edipi_entry.pack(side="left", padx=(8, 6))
        self._edipi_entry.bind("<Return>", lambda _: self._do_lookup())

        self._lookup_btn = tk.Button(
            edipi_row, text="Look Up", font=FBOLD,
            bg=C["accent"], fg="white",
            activebackground=C["accent_light"],
            relief="flat", padx=12, pady=4, cursor="hand2",
            command=self._do_lookup
        )
        self._lookup_btn.pack(side="left")

        # Result panel (hidden until lookup fires)
        self._result_frame = tk.Frame(self._lookup_frame, bg=C["surface"],
                                      highlightbackground=C["border"],
                                      highlightthickness=1)
        # Not packed yet — shown after successful lookup

        self._result_lbl = tk.Label(
            self._result_frame, text="", font=("Arial", 9),
            fg=C["text"], bg=C["surface"],
            justify="left", anchor="w", padx=12, pady=8
        )
        self._result_lbl.pack(fill="x")

        confirm_row = tk.Frame(self._result_frame, bg=C["surface"])
        confirm_row.pack(fill="x", padx=12, pady=(0, 10))
        self._confirm_btn = tk.Button(
            confirm_row, text="✔  Link FASC-N to this record & Log",
            font=FBOLD, bg=C["green"], fg="white",
            activebackground="#00A368", relief="flat",
            padx=12, pady=5, cursor="hand2",
            command=self._confirm_link
        )
        self._confirm_btn.pack(side="left", padx=(0, 8))
        tk.Button(
            confirm_row, text="Not this person",
            font=FSML, bg=C["surface2"], fg=C["dim"],
            relief="flat", padx=10, pady=5, cursor="hand2",
            command=self._clear_result
        ).pack(side="left")

        # "Not in database" switch
        switch_row = tk.Frame(self._lookup_frame, bg=C["bg"])
        switch_row.pack(fill="x", pady=(12, 0))
        tk.Label(switch_row, text="Person not in database yet?",
                 font=("Arial", 9), fg=C["dim"], bg=C["bg"]
                 ).pack(side="left")
        tk.Button(
            switch_row, text="Enter details manually →",
            font=("Arial", 9, "bold"), fg=C["gold"], bg=C["bg"],
            relief="flat", cursor="hand2", bd=0,
            command=self._show_manual_mode
        ).pack(side="left", padx=4)

    def _build_manual_section(self):
        """Full manual-entry form — shown in manual mode."""
        self._manual_frame = tk.Frame(self, bg=C["bg"], padx=16, pady=10)

        tk.Label(self._manual_frame,
                 text="Enter personnel details to create a new record.",
                 font=("Arial", 9), fg=C["dim"], bg=C["bg"]
                 ).pack(anchor="w", pady=(0, 8))

        # Back to lookup link
        back_row = tk.Frame(self._manual_frame, bg=C["bg"])
        back_row.pack(fill="x", pady=(0, 6))
        tk.Button(
            back_row, text="← Back to EDIPI lookup",
            font=("Arial", 9), fg=C["gold"], bg=C["bg"],
            relief="flat", cursor="hand2", bd=0,
            command=self._show_lookup_mode
        ).pack(side="left")

        body = tk.Frame(self._manual_frame, bg=C["bg"])
        body.pack(fill="both")

        self._vars = {}
        fields = [
            ("Last Name *",    "last_name",   "entry", None),
            ("First Name *",   "first_name",  "entry", None),
            ("Middle Name",    "middle_name", "entry", None),
            ("DoD ID (EDIPI)", "edipi",       "entry", None),
            ("Date of Birth",  "dob",         "entry", None),
            ("Gender",         "gender",      "combo", self.GENDERS),
            ("Branch",         "branch",      "combo", self.BRANCHES),
            ("Rank / Grade",   "rank",        "entry", None),
            ("Affiliation",    "affiliation", "entry", None),
        ]
        self._first_manual_entry = None
        for row_idx, (label, key, wtype, opts) in enumerate(fields):
            tk.Label(body, text=label, font=FSML, fg=C["dim"], bg=C["bg"],
                     anchor="w", width=16
                     ).grid(row=row_idx, column=0, sticky="w", pady=2)
            var = tk.StringVar()
            self._vars[key] = var
            if wtype == "entry":
                e = tk.Entry(body, textvariable=var, font=FMONO10,
                             bg=C["surface2"], fg=C["text"],
                             insertbackground=C["text"], relief="flat",
                             highlightbackground=C["border"],
                             highlightthickness=1, width=28)
                e.grid(row=row_idx, column=1, sticky="ew", padx=(8,0), pady=2)
                if self._first_manual_entry is None:
                    self._first_manual_entry = e
            else:
                cb = ttk.Combobox(body, textvariable=var, values=opts,
                                  state="readonly", width=26, font=FSML)
                cb.grid(row=row_idx, column=1, sticky="ew", padx=(8,0), pady=2)

        tk.Label(body, text="* Required", font=("Arial", 8), fg=C["dim"],
                 bg=C["bg"]).grid(row=len(fields), column=0, columnspan=2,
                                  sticky="w", pady=(4,0))

    def _build_direction_row(self):
        self._dir_frame = tk.Frame(self, bg=C["bg"], padx=16, pady=8)
        tk.Label(self._dir_frame, text="Log as:", font=FSML,
                 fg=C["dim"], bg=C["bg"]).pack(side="left")
        self._dir_var = tk.StringVar(value="OUT")
        for val, lbl, col in [("OUT","CHECK OUT",C["red"]),
                               ("IN", "CHECK IN", C["green"])]:
            tk.Radiobutton(
                self._dir_frame, text=lbl, variable=self._dir_var,
                value=val, font=("Arial", 10, "bold"), fg=col, bg=C["bg"],
                selectcolor=C["surface"], activebackground=C["bg"]
            ).pack(side="left", padx=12)

    def _build_buttons(self):
        self._btn_frame = tk.Frame(self, bg=C["surface"], pady=10)
        self._btn_frame.pack(fill="x", side="bottom")

        tk.Button(self._btn_frame, text="Cancel", font=FSML,
                  bg=C["surface2"], fg=C["dim"], relief="flat",
                  padx=16, pady=6, cursor="hand2",
                  command=self._cancel
                  ).pack(side="right", padx=8)

        # Save button — only visible in manual mode
        self._save_btn = tk.Button(
            self._btn_frame, text="Save & Log",
            font=("Arial", 10, "bold"),
            bg=C["accent"], fg="white",
            activebackground=C["accent_light"],
            relief="flat", padx=16, pady=6, cursor="hand2",
            command=self._save_manual
        )

    # ── Mode switching ────────────────────────────────────────────────────────

    def _show_lookup_mode(self):
        self._manual_frame.pack_forget()
        self._save_btn.pack_forget()
        self._dir_frame.pack_forget()

        self._lookup_frame.pack(fill="both", padx=0, pady=0, before=self._btn_frame)
        self._dir_frame.pack(fill="x", before=self._btn_frame)
        self._mode_lbl.config(text="Step 1: EDIPI lookup")
        self._edipi_entry.focus_set()
        self.after(60, self._center)

    def _show_manual_mode(self):
        self._lookup_frame.pack_forget()
        self._save_btn.pack_forget()

        self._manual_frame.pack(fill="both", padx=0, pady=0, before=self._dir_frame)
        self._save_btn.pack(side="right", padx=4, in_=self._btn_frame)
        self._mode_lbl.config(text="Step 2: Manual entry")
        if self._first_manual_entry:
            self._first_manual_entry.focus_set()
        self.after(60, self._center)

    # ── Lookup logic ──────────────────────────────────────────────────────────

    def _do_lookup(self):
        edipi = self._edipi_var.get().strip()
        edipi = re.sub(r"\D", "", edipi)   # digits only
        if not edipi:
            messagebox.showwarning("Enter EDIPI",
                                   "Type a DoD ID (EDIPI) number first.",
                                   parent=self)
            return

        person = self.db.lookup_by_edipi(edipi)
        if not person:
            # Also try EDIPI: key format
            person = self.db.lookup_exact(f"EDIPI:{edipi}")

        if not person:
            messagebox.showinfo(
                "Not Found",
                f"No record found for EDIPI {edipi}.\n\n"
                "Click \"Enter details manually\" to create a new record.",
                parent=self
            )
            return

        self._matched = person
        name  = f"{person.get('first_name','')} {person.get('last_name','')}".strip()
        dob   = person.get("dob", "") or "—"
        br    = person.get("branch", "") or "—"
        rank  = person.get("rank", "") or "—"
        cur_key = person.get("fascn_key", "")

        self._result_lbl.config(
            text=f"  ✔  Record found:\n\n"
                 f"      Name:    {name}\n"
                 f"      EDIPI:   {edipi}\n"
                 f"      DOB:     {dob}\n"
                 f"      Branch:  {br}   Rank: {rank}\n"
                 f"      Current key: {cur_key}\n\n"
                 f"  Link this FASC-N to that record?"
        )
        self._result_frame.pack(fill="x", pady=(10, 0))
        self.after(60, self._center)

    def _clear_result(self):
        self._matched = None
        self._result_frame.pack_forget()
        self._edipi_var.set("")
        self._edipi_entry.focus_set()
        self.after(60, self._center)

    def _confirm_link(self):
        """Link the scanned FASC-N to the found record and log the event."""
        if not self._matched:
            return
        direction = self._dir_var.get()
        self.destroy()
        self.callback(
            dict(self._matched),
            direction,
            link_fascn=self.fascn_key
        )

    # ── Manual save ───────────────────────────────────────────────────────────

    def _save_manual(self):
        data = {k: v.get().strip() for k, v in self._vars.items()}
        if not data["last_name"] or not data["first_name"]:
            messagebox.showwarning("Required Fields",
                                   "Last Name and First Name are required.",
                                   parent=self)
            return
        data["raw_fascn"] = self.fascn_key
        self.destroy()
        self.callback(data, self._dir_var.get(), link_fascn=None)

    def _cancel(self):
        self.destroy()


# =============================================================================
# PERSON EDIT MODAL
# =============================================================================

class PersonEditModal(tk.Toplevel):
    """Add a new record or edit an existing one."""

    BRANCHES = ["","ARMY","NAVY","USMC","USAF","USSF","USCG",
                "CIVILIAN","CONTRACTOR","OTHER"]
    GENDERS  = ["","M","F","X"]

    def __init__(self, parent, db: RosterDB, existing: dict = None, callback=None):
        super().__init__(parent)
        self.db       = db
        self.existing = existing   # None = new record
        self.callback = callback
        self.title("Edit Record" if existing else "Add New Record")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self._build()
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.after(50, self._center)

    def _center(self):
        self.update_idletasks()
        pw = self.master.winfo_rootx()
        py = self.master.winfo_rooty()
        w  = self.winfo_width()
        h  = self.winfo_height()
        self.geometry(
            f"+{pw + self.master.winfo_width()//2 - w//2}"
            f"+{py + self.master.winfo_height()//2 - h//2}"
        )

    def _build(self):
        hdr = tk.Frame(self, bg=C["accent"], pady=8)
        hdr.pack(fill="x")
        lbl = "Edit Personnel Record" if self.existing else "Add New Personnel Record"
        tk.Label(hdr, text=lbl, font=("Arial",11,"bold"),
                 fg="white", bg=C["accent"]).pack(padx=16)

        body = tk.Frame(self, bg=C["bg"], padx=16, pady=12)
        body.pack(fill="both")

        self._vars = {}
        ex = self.existing or {}

        fields = [
            ("FASC-N / Key *",  "fascn_key",      "entry",  None,         ex.get("fascn_key","")),
            ("DoD ID (EDIPI)",  "edipi",           "entry",  None,         ex.get("edipi","")),
            ("Last Name *",     "last_name",       "entry",  None,         ex.get("last_name","")),
            ("First Name *",    "first_name",      "entry",  None,         ex.get("first_name","")),
            ("Middle Name",     "middle_name",     "entry",  None,         ex.get("middle_name","")),
            ("Unit / Department","unit",           "entry",  None,         ex.get("unit","")),
            ("Gender",          "gender",          "combo",  self.GENDERS, ex.get("gender","")),
            ("Branch",          "branch",          "combo",  self.BRANCHES,ex.get("branch","")),
            ("Rank / Grade",    "rank",            "entry",  None,         ex.get("rank","")),
            ("Affiliation",     "affiliation",     "entry",  None,         ex.get("affiliation","")),
        ]

        first_entry = None
        for row_idx, (label, key, wtype, opts, val) in enumerate(fields):
            tk.Label(body, text=label, font=FSML, fg=C["dim"], bg=C["bg"],
                     anchor="w", width=17).grid(row=row_idx, column=0,
                                                sticky="w", pady=2)
            var = tk.StringVar(value=val)
            self._vars[key] = var
            if wtype == "entry":
                show = ""
                e = tk.Entry(body, textvariable=var, font=FMONO10,
                             bg=C["surface2"], fg=C["text"],
                             insertbackground=C["text"], relief="flat",
                             highlightbackground=C["border"],
                             highlightthickness=1, width=30, show=show)
                e.grid(row=row_idx, column=1, sticky="ew", padx=(8,0), pady=2)
                if key == "fascn_key" and self.existing:
                    e.config(state="disabled")   # can't change PK
                if first_entry is None and key not in ("fascn_key",):
                    first_entry = e
            else:
                cb = ttk.Combobox(body, textvariable=var, values=opts,
                                  state="readonly", width=28, font=FSML)
                cb.grid(row=row_idx, column=1, sticky="ew", padx=(8,0), pady=2)

        tk.Label(body, text="* Required",
                 font=("Arial",8), fg=C["dim"], bg=C["bg"]
                 ).grid(row=len(fields), column=0, columnspan=2,
                        sticky="w", pady=(4,0))

        if first_entry:
            first_entry.focus_set()

        btn_row = tk.Frame(self, bg=C["surface"], pady=8)
        btn_row.pack(fill="x")
        tk.Button(btn_row, text="Cancel", font=FSML,
                  bg=C["surface2"], fg=C["dim"], relief="flat",
                  padx=14, pady=5, cursor="hand2",
                  command=self.destroy).pack(side="right", padx=8)
        tk.Button(btn_row, text="Save Record", font=("Arial",10,"bold"),
                  bg=C["accent"], fg="white",
                  activebackground=C["accent_light"],
                  relief="flat", padx=14, pady=5, cursor="hand2",
                  command=self._save).pack(side="right", padx=4)

    def _save(self):
        data = {k: v.get().strip() for k, v in self._vars.items()}
        if not data.get("last_name") or not data.get("first_name"):
            messagebox.showwarning("Required", "Last Name and First Name are required.",
                                   parent=self)
            return
        key = data.get("fascn_key", "").strip()
        if not key:
            messagebox.showwarning("Required", "FASC-N / Key is required.", parent=self)
            return
        if self.existing:
            data.setdefault("dob", (self.existing or {}).get("dob", ""))
            data.setdefault("place_of_birth", (self.existing or {}).get("place_of_birth", ""))
            self.db.update_person(key, data, performed_by="operator")
        else:
            self.db.upsert_person(key, data, added_by="manual",
                                  performed_by="operator")
        self.db.update_unit(key, data.get("unit", ""))
        self.destroy()
        if self.callback:
            self.callback()


# =============================================================================
# ROSTER MANAGEMENT WINDOW
# =============================================================================

class RosterWindow(tk.Toplevel):
    def __init__(self, parent, db: RosterDB, refresh_cb):
        super().__init__(parent)
        self.title("Roster Management")
        self.configure(bg=C["bg"])
        self.db = db
        self.refresh_cb = refresh_cb
        self.geometry("980x620")
        self.transient(parent)
        self._build()
        self._load_roster()

    def _build(self):
        # ── Header bar ──────────────────────────────────────────────────
        top = tk.Frame(self, bg=C["surface"], pady=8)
        top.pack(fill="x")
        tk.Label(top, text="ROSTER MANAGEMENT", font=("Arial",11,"bold"),
                 fg=C["gold"], bg=C["surface"]).pack(side="left", padx=16)
        for lbl, cmd, bg in [
            ("+ Add Record",       self._add,         C["green"]),
            ("Import Roster",      self._import_csv,  C["accent"]),
            ("Template CSV",       self._dl_template, C["surface2"]),
            ("Export Audit Log",   self._export_audit, C["surface2"]),
        ]:
            tk.Button(top, text=lbl, font=FSML, bg=bg,
                      fg="white" if bg not in (C["surface2"],) else C["dim"],
                      relief="flat", padx=10, pady=4, cursor="hand2",
                      command=cmd).pack(side="right", padx=4)

        # ── Notebook (three tabs) ─────────────────────────────────────────
        nb_style = ttk.Style()
        nb_style.configure("TNotebook", background=C["bg"], borderwidth=0)
        nb_style.configure("TNotebook.Tab", background=C["surface2"],
                           foreground=C["dim"], padding=[12,4],
                           font=("Arial",9,"bold"))
        nb_style.map("TNotebook.Tab",
                     background=[("selected", C["accent"])],
                     foreground=[("selected", "white")])

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=10, pady=8)

        roster_tab  = tk.Frame(self.nb, bg=C["bg"])
        audit_tab   = tk.Frame(self.nb, bg=C["bg"])
        lr_log_tab  = tk.Frame(self.nb, bg=C["bg"])
        leave_tab   = tk.Frame(self.nb, bg=C["bg"])
        self.nb.add(roster_tab,  text="  Personnel Roster  ")
        self.nb.add(audit_tab,   text="  Roster Audit Log  ")
        self.nb.add(lr_log_tab,  text="  ⚠ Liberty Risk Log  ")
        self.nb.add(leave_tab,   text="  🏖 Leave Status  ")

        self._build_roster_tab(roster_tab)
        self._build_audit_tab(audit_tab)
        self._build_lr_log_tab(lr_log_tab)
        self._build_roster_leave_tab(leave_tab)

        # Status bar
        self.status_lbl = tk.Label(self, text="", font=FSML,
                                    fg=C["dim"], bg=C["bg"])
        self.status_lbl.pack(anchor="w", padx=14, pady=(0,6))

    # ── Roster tab ──────────────────────────────────────────────────────

    def _build_roster_tab(self, parent):
        # Filter + action bar
        fb = tk.Frame(parent, bg=C["bg"], pady=6)
        fb.pack(fill="x", padx=8)
        tk.Label(fb, text="Search:", font=FSML, fg=C["dim"],
                 bg=C["bg"]).pack(side="left")
        self._fvar = tk.StringVar()
        tk.Entry(fb, textvariable=self._fvar, font=FSML,
                 bg=C["surface2"], fg=C["text"], relief="flat",
                 insertbackground=C["text"]
                 ).pack(side="left", padx=6, fill="x", expand=True)
        self._fvar.trace_add("write", lambda *_: self._load_roster())
        tk.Button(fb, text="Edit Selected", font=FSML,
                  bg=C["accent"], fg="white", relief="flat",
                  padx=10, pady=3, cursor="hand2",
                  command=self._edit_selected).pack(side="right", padx=4)
        tk.Button(fb, text="Delete Selected", font=FSML,
                  bg=C["red"], fg="white", relief="flat",
                  padx=10, pady=3, cursor="hand2",
                  command=self._delete_selected).pack(side="right", padx=4)
        tk.Button(fb, text="⚠ Toggle Liberty Risk", font=FSML,
                  bg="#CC6600", fg="white", relief="flat",
                  padx=10, pady=3, cursor="hand2",
                  command=self._toggle_liberty_risk).pack(side="right", padx=4)

        # Treeview
        cols = ("FASC-N","EDIPI","Last Name","First Name",
                "DOB","Branch","Rank","LR","Added By","Added At")
        tf = tk.Frame(parent, bg=C["bg"])
        tf.pack(fill="both", expand=True, padx=8, pady=(0,4))

        style = ttk.Style()
        style.configure("Roster.Treeview",
                        background=C["surface"], foreground=C["text"],
                        fieldbackground=C["surface"], rowheight=24,
                        font=FMONO10)
        style.configure("Roster.Treeview.Heading",
                        background=C["surface2"], foreground=C["gold"],
                        font=("Arial",9,"bold"), relief="flat")
        style.map("Roster.Treeview",
                  background=[("selected", C["accent"])],
                  foreground=[("selected","white")])

        self.r_tree = ttk.Treeview(tf, columns=cols, show="headings",
                                    style="Roster.Treeview", selectmode="browse")
        widths = [155,100,110,110,90,70,60,40,85,130]
        for col, w in zip(cols, widths):
            self.r_tree.heading(col, text=col,
                                command=lambda c=col: self._sort_roster(c))
            self.r_tree.column(col, width=w, minwidth=40)
        self.r_tree.bind("<Double-1>", lambda _: self._edit_selected())

        # Liberty Risk row tag — bright orange/red background
        self.r_tree.tag_configure("liberty_risk",
                                   background="#4A1000",
                                   foreground="#FF8844")

        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.r_tree.yview)
        self.r_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.r_tree.pack(fill="both", expand=True)
        self._r_sort_col, self._r_sort_rev = None, False

    # ── Audit tab ───────────────────────────────────────────────────────

    def _build_audit_tab(self, parent):
        fb = tk.Frame(parent, bg=C["bg"], pady=6)
        fb.pack(fill="x", padx=8)
        tk.Label(fb, text="Filter:", font=FSML, fg=C["dim"],
                 bg=C["bg"]).pack(side="left")
        self._afvar = tk.StringVar()
        tk.Entry(fb, textvariable=self._afvar, font=FSML,
                 bg=C["surface2"], fg=C["text"], relief="flat",
                 insertbackground=C["text"]
                 ).pack(side="left", padx=6, fill="x", expand=True)
        self._afvar.trace_add("write", lambda *_: self._load_audit())

        cols = ("Timestamp","Action","Last Name","First Name",
                "EDIPI","Changed Fields","Old Values","New Values","By")
        tf = tk.Frame(parent, bg=C["bg"])
        tf.pack(fill="both", expand=True, padx=8, pady=(0,4))

        style = ttk.Style()
        style.configure("Audit.Treeview",
                        background=C["surface"], foreground=C["text"],
                        fieldbackground=C["surface"], rowheight=24,
                        font=FMONO10)
        style.configure("Audit.Treeview.Heading",
                        background=C["surface2"], foreground=C["gold"],
                        font=("Arial",9,"bold"), relief="flat")
        style.map("Audit.Treeview",
                  background=[("selected", C["accent"])],
                  foreground=[("selected","white")])

        self.a_tree = ttk.Treeview(tf, columns=cols, show="headings",
                                    style="Audit.Treeview", selectmode="browse")
        widths = [150,80,110,110,100,130,160,160,80]
        for col, w in zip(cols, widths):
            self.a_tree.heading(col, text=col)
            self.a_tree.column(col, width=w, minwidth=50)

        # Colour rows by action
        self.a_tree.tag_configure("INSERT",     background="#001F3F", foreground="#7FC8FF")
        self.a_tree.tag_configure("UPDATE",     background="#1A1A00", foreground="#FFE066")
        self.a_tree.tag_configure("DELETE",     background=C["out_bg"], foreground=C["out_fg"])
        self.a_tree.tag_configure("FASCN_LINK", background="#1A0030", foreground="#CC99FF")
        self.a_tree.tag_configure("CSV_IMPORT", background="#001A00", foreground="#66FF99")

        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.a_tree.yview)
        self.a_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.a_tree.pack(fill="both", expand=True)

    # ── Data loaders ────────────────────────────────────────────────────

    def _load_roster(self):
        q = self._fvar.get().lower()
        for r in self.r_tree.get_children():
            self.r_tree.delete(r)
        cur = self.db.conn.execute(
            "SELECT fascn_key, edipi, last_name, first_name, "
            "dob, branch, rank, liberty_risk, added_by, added_at "
            "FROM personnel ORDER BY last_name, first_name"
        )
        count = 0
        for row in cur.fetchall():
            vals = list(row)
            if q and not any(q in str(v).lower() for v in vals):
                continue
            is_lr   = bool(vals[7])
            lr_flag = "⚠ LR" if is_lr else ""
            # Replace the integer with display flag
            display = vals[:7] + [lr_flag] + vals[8:]
            tag = "liberty_risk" if is_lr else ""
            self.r_tree.insert("", "end", values=display,
                               iid=vals[0], tags=(tag,) if tag else ())
            count += 1
        self._update_status(count)

    def _load_audit(self):
        q = self._afvar.get()
        rows = self.db.get_audit_log(text_filter=q)
        for r in self.a_tree.get_children():
            self.a_tree.delete(r)
        for row in rows:
            action = row.get("action","")
            tag = (action if action in
                   ("INSERT","UPDATE","DELETE","FASCN_LINK") else "CSV_IMPORT")
            self.a_tree.insert("", "end", values=(
                row["timestamp"], action,
                row["last_name"], row["first_name"], row["edipi"],
                row["changed_fields"], row["old_values"],
                row["new_values"], row["performed_by"],
            ), tags=(tag,))
        n = self.db.audit_count()
        self.status_lbl.config(
            text=f"{len(rows)} audit entries shown / {n} total")

    def _update_status(self, shown: int):
        total = self.db.roster_count()
        self.status_lbl.config(
            text=f"{shown} records shown / {total} total  |  "
                 f"Audit entries: {self.db.audit_count()}")

    # ── Roster actions ───────────────────────────────────────────────────

    def _add(self):
        PersonEditModal(self, self.db, existing=None,
                        callback=self._after_edit)

    def _edit_selected(self):
        sel = self.r_tree.selection()
        if not sel:
            messagebox.showinfo("No Selection", "Select a record to edit.",
                                parent=self)
            return
        fascn_key = sel[0]
        existing  = self.db.lookup_exact(fascn_key)
        if not existing:
            messagebox.showerror("Not Found",
                                 "Record no longer exists.", parent=self)
            return
        PersonEditModal(self, self.db, existing=existing,
                        callback=self._after_edit)

    def _delete_selected(self):
        sel = self.r_tree.selection()
        if not sel:
            messagebox.showinfo("No Selection", "Select a record to delete.",
                                parent=self)
            return
        fascn_key = sel[0]
        existing  = self.db.lookup_exact(fascn_key)
        name = (f"{existing.get('first_name','')} "
                f"{existing.get('last_name','')}").strip() if existing else fascn_key
        if not messagebox.askyesno(
                "Confirm Delete",
                f"Permanently remove {name} from the roster?\n\n"
                f"FASC-N: {fascn_key}\n\n"
                "This action will be recorded in the audit log.",
                icon="warning", parent=self):
            return
        self.db.delete_person(fascn_key, performed_by="operator")
        self._after_edit()
        messagebox.showinfo("Deleted",
                            f"{name} removed from roster.", parent=self)

    def _after_edit(self):
        self._load_roster()
        self._load_audit()
        self._load_lr_log()
        self._load_roster_leave()
        self.refresh_cb()

    def _toggle_liberty_risk(self):
        sel = self.r_tree.selection()
        if not sel:
            messagebox.showinfo("No Selection",
                                "Select a person to toggle Liberty Risk status.",
                                parent=self)
            return
        fascn_key = sel[0]
        existing  = self.db.lookup_exact(fascn_key)
        if not existing:
            return
        name = (f"{existing.get('first_name','')} "
                f"{existing.get('last_name','')}").strip()
        currently = bool(existing.get("liberty_risk", 0))

        if currently:
            # Clear the flag
            if messagebox.askyesno(
                "Clear Liberty Risk",
                f"Remove Liberty Risk designation from {name}?\n\n"
                "This action will be recorded in the audit log.",
                parent=self
            ):
                self.db.set_liberty_risk(fascn_key, False, "",
                                          performed_by="operator")
                self._after_edit()
                messagebox.showinfo("Updated",
                                    f"{name} Liberty Risk flag cleared.",
                                    parent=self)
        else:
            # Set the flag — prompt for reason
            dlg = tk.Toplevel(self)
            dlg.title("Set Liberty Risk")
            dlg.configure(bg=C["bg"])
            dlg.resizable(False, False)
            dlg.grab_set()
            dlg.transient(self)

            tk.Frame(dlg, bg="#CC0000", height=4).pack(fill="x")
            tk.Label(dlg,
                     text=f"Set Liberty Risk — {name}",
                     font=("Arial", 11, "bold"),
                     fg="#FF4444", bg=C["surface"], pady=10
                     ).pack(fill="x", padx=16)

            body = tk.Frame(dlg, bg=C["bg"], padx=16, pady=12)
            body.pack(fill="both")
            tk.Label(body,
                     text="Reason for Liberty Risk designation *",
                     font=FBOLD, fg=C["text"], bg=C["bg"]
                     ).pack(anchor="w")
            reason_text = tk.Text(body, height=4, width=44,
                                   font=FSML,
                                   bg=C["surface2"], fg=C["text"],
                                   insertbackground=C["text"],
                                   relief="flat",
                                   highlightbackground=C["border"],
                                   highlightthickness=1)
            reason_text.pack(fill="x", pady=(4, 0))
            tk.Label(body, text="* Required — recorded in audit log",
                     font=("Arial", 8), fg=C["dim"], bg=C["bg"]
                     ).pack(anchor="w", pady=(2, 0))

            err_lbl = tk.Label(body, text="", font=("Arial", 9),
                                fg=C["red"], bg=C["bg"])
            err_lbl.pack(anchor="w")

            def _save():
                reason = reason_text.get("1.0", "end").strip()
                if not reason:
                    err_lbl.config(text="⚠  A reason is required.")
                    return
                self.db.set_liberty_risk(fascn_key, True, reason,
                                          performed_by="operator")
                dlg.destroy()
                self._after_edit()
                messagebox.showinfo("Updated",
                                    f"{name} flagged as Liberty Risk.",
                                    parent=self)

            btn_row = tk.Frame(dlg, bg=C["surface"], pady=8)
            btn_row.pack(fill="x")
            tk.Button(btn_row, text="Cancel", font=FSML,
                      bg=C["surface2"], fg=C["dim"], relief="flat",
                      padx=14, pady=5, cursor="hand2",
                      command=dlg.destroy
                      ).pack(side="right", padx=8)
            tk.Button(btn_row, text="⚠  Flag as Liberty Risk",
                      font=("Arial", 10, "bold"),
                      bg="#CC0000", fg="white",
                      activebackground="#AA0000",
                      relief="flat", padx=14, pady=5, cursor="hand2",
                      command=_save
                      ).pack(side="right", padx=4)

    def _build_roster_leave_tab(self, parent):
        fb = tk.Frame(parent, bg=C["bg"], pady=6)
        fb.pack(fill="x", padx=8)
        tk.Label(fb, text="Filter:", font=FSML, fg=C["dim"],
                 bg=C["bg"]).pack(side="left")
        self._rwl_fvar = tk.StringVar()
        tk.Entry(fb, textvariable=self._rwl_fvar, font=FSML,
                 bg=C["surface2"], fg=C["text"], relief="flat",
                 insertbackground=C["text"]
                 ).pack(side="left", padx=6, fill="x", expand=True)
        self._rwl_fvar.trace_add("write", lambda *_: self._load_roster_leave())
        self._rwl_status = tk.StringVar(value="ALL")
        for val, lbl in [("ALL", "All"), ("ACTIVE", "On Leave"),
                         ("RETURNED", "Returned")]:
            tk.Radiobutton(fb, text=lbl, variable=self._rwl_status, value=val,
                           font=FSML, fg=C["dim"], bg=C["bg"],
                           selectcolor=C["surface"], activebackground=C["bg"],
                           command=self._load_roster_leave
                           ).pack(side="left", padx=3)

        ab = tk.Frame(parent, bg=C["bg"])
        ab.pack(fill="x", padx=8, pady=(0, 4))
        tk.Button(ab, text="Mark Returned", font=FSML, bg=C["green"], fg="white",
                  relief="flat", padx=10, pady=3, cursor="hand2",
                  command=self._rw_leave_mark_returned).pack(side="left", padx=3)
        tk.Button(ab, text="Reopen Leave", font=FSML, bg="#8A6000", fg="white",
                  relief="flat", padx=10, pady=3, cursor="hand2",
                  command=self._rw_leave_reopen).pack(side="left", padx=3)
        tk.Button(ab, text="Edit Dates", font=FSML, bg=C["accent"], fg="white",
                  relief="flat", padx=10, pady=3, cursor="hand2",
                  command=self._rw_leave_edit).pack(side="left", padx=3)
        tk.Button(ab, text="Delete Record", font=FSML, bg=C["red"], fg="white",
                  relief="flat", padx=10, pady=3, cursor="hand2",
                  command=self._rw_leave_delete).pack(side="left", padx=3)
        tk.Button(ab, text="Export CSV", font=FSML, bg=C["surface2"], fg=C["dim"],
                  relief="flat", padx=10, pady=3, cursor="hand2",
                  command=self._rw_leave_export).pack(side="right", padx=3)

        tf = tk.Frame(parent, bg=C["bg"])
        tf.pack(fill="both", expand=True, padx=8, pady=(0, 4))
        style = ttk.Style()
        style.configure("RWLeave.Treeview",
                        background=C["surface"], foreground=C["text"],
                        fieldbackground=C["surface"], rowheight=24, font=FMONO10)
        style.configure("RWLeave.Treeview.Heading",
                        background=C["surface2"], foreground=C["gold"],
                        font=("Arial", 9, "bold"), relief="flat")
        style.map("RWLeave.Treeview",
                  background=[("selected", C["accent"])],
                  foreground=[("selected", "white")])
        cols = ("Status", "Type", "Last Name", "First Name", "EDIPI",
                "Start Date", "End Date", "Departed", "Returned")
        self.rwl_tree = ttk.Treeview(tf, columns=cols, show="headings",
                                     style="RWLeave.Treeview", selectmode="browse")
        widths = [70, 120, 110, 110, 100, 95, 95, 150, 150]
        for col, w in zip(cols, widths):
            self.rwl_tree.heading(col, text=col)
            self.rwl_tree.column(col, width=w, minwidth=50)
        self.rwl_tree.tag_configure("ACTIVE",
                                    background="#2A2000", foreground=C["gold"])
        self.rwl_tree.tag_configure("RETURNED",
                                    background=C["surface"], foreground=C["dim"])
        self.rwl_tree.bind("<Double-1>", lambda _: self._rw_leave_edit())
        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.rwl_tree.yview)
        self.rwl_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.rwl_tree.pack(fill="both", expand=True)
        self.rwl_count_lbl = tk.Label(parent, text="", font=FSML,
                                      fg=C["dim"], bg=C["bg"])
        self.rwl_count_lbl.pack(anchor="e", padx=14, pady=(0, 4))

    def _load_roster_leave(self):
        if not hasattr(self, "rwl_tree"):
            return
        txt = self._rwl_fvar.get() if hasattr(self, "_rwl_fvar") else ""
        status = self._rwl_status.get() if hasattr(self, "_rwl_status") else "ALL"
        rows = self.db.get_leave_records(status_filter=status, text_filter=txt)
        for r in self.rwl_tree.get_children():
            self.rwl_tree.delete(r)
        for row in rows:
            dep = (row.get("departure_ts", "") or "").replace("T", " ")
            ret = (row.get("return_ts", "") or "").replace("T", " ") or "-"
            self.rwl_tree.insert("", "end", iid=str(row["id"]), values=(
                row.get("status", ""), row.get("leave_type", ""),
                row.get("last_name", ""), row.get("first_name", ""),
                row.get("edipi", ""), row.get("start_date", ""),
                row.get("end_date", ""), dep, ret,
            ), tags=(row.get("status", ""),))
        total = self.db.leave_total_count()
        active = self.db.leave_active_count()
        self.rwl_count_lbl.config(
            text=f"{len(rows)} shown / {total} total  |  "
                 f"{active} currently on leave")

    def _rw_leave_selected_id(self):
        sel = self.rwl_tree.selection()
        if not sel:
            messagebox.showinfo("No Selection",
                                "Select a leave record first.", parent=self)
            return None
        return int(sel[0])

    def _rw_leave_after(self):
        self._load_roster_leave()
        if hasattr(self.master, "_refresh_leave_log"):
            try:
                self.master._refresh_leave_log()
            except Exception:
                pass

    def _rw_leave_mark_returned(self):
        lid = self._rw_leave_selected_id()
        if lid is None:
            return
        rec = self.db.get_leave_record(lid)
        if not rec:
            return
        if rec["status"] != "ACTIVE":
            messagebox.showinfo("Already Returned",
                "This person is already marked as returned.", parent=self)
            return
        name = f"{rec.get('first_name','')} {rec.get('last_name','')}".strip()
        if messagebox.askyesno("Mark Returned",
                f"Mark {name} as returned from {rec['leave_type']}?", parent=self):
            self.db.mark_leave_returned(lid)
            self._rw_leave_after()

    def _rw_leave_reopen(self):
        lid = self._rw_leave_selected_id()
        if lid is None:
            return
        rec = self.db.get_leave_record(lid)
        if not rec:
            return
        if rec["status"] == "ACTIVE":
            messagebox.showinfo("Already Active",
                "This leave is already active.", parent=self)
            return
        name = f"{rec.get('first_name','')} {rec.get('last_name','')}".strip()
        if messagebox.askyesno("Reopen Leave",
                f"Reopen {rec['leave_type']} for {name}? "
                "They will be considered on leave again.", parent=self):
            self.db.reopen_leave(lid)
            self._rw_leave_after()

    def _rw_leave_edit(self):
        lid = self._rw_leave_selected_id()
        if lid is None:
            return
        rec = self.db.get_leave_record(lid)
        if not rec:
            return
        def _on_save(start_iso, end_iso):
            if not start_iso:
                return
            self.db.update_leave_dates(lid, start_iso, end_iso)
            self._rw_leave_after()
        LeaveEditModal(self, rec, _on_save)

    def _rw_leave_delete(self):
        lid = self._rw_leave_selected_id()
        if lid is None:
            return
        rec = self.db.get_leave_record(lid)
        if not rec:
            return
        name = f"{rec.get('first_name','')} {rec.get('last_name','')}".strip()
        if messagebox.askyesno("Delete Leave Record",
                f"Permanently delete this {rec['leave_type']} record for {name} "
                f"({rec['start_date']} to {rec['end_date']})? This cannot be undone.",
                icon="warning", parent=self):
            self.db.delete_leave_record(lid)
            self._rw_leave_after()

    def _rw_leave_export(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV", "*.csv"), ("All", "*.*")],
            initialfile=f"leave_records_{datetime.date.today()}.csv", parent=self)
        if not path:
            return
        rows = self.db.get_leave_records()
        cols = ["status", "leave_type", "last_name", "first_name", "edipi",
                "start_date", "end_date", "departure_ts", "return_ts", "fascn_key"]
        with open(path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)
        messagebox.showinfo("Exported",
                            f"Leave records saved: {path}", parent=self)

    def _build_lr_log_tab(self, parent):
        fb = tk.Frame(parent, bg=C["bg"], pady=6)
        fb.pack(fill="x", padx=8)
        tk.Label(fb, text="Filter:", font=FSML, fg=C["dim"],
                 bg=C["bg"]).pack(side="left")
        self._lrfvar = tk.StringVar()
        tk.Entry(fb, textvariable=self._lrfvar, font=FSML,
                 bg=C["surface2"], fg=C["text"], relief="flat",
                 insertbackground=C["text"]
                 ).pack(side="left", padx=6, fill="x", expand=True)
        self._lrfvar.trace_add("write", lambda *_: self._load_lr_log())

        tk.Button(fb, text="Export CSV", font=FSML,
                  bg=C["accent"], fg="white", relief="flat",
                  padx=10, pady=3, cursor="hand2",
                  command=self._export_lr_log
                  ).pack(side="right", padx=4)

        cols = ("Timestamp","Last Name","First Name","EDIPI",
                "Direction","Destination","Acknowledger Name",
                "Acknowledger EDIPI","Reason")
        tf = tk.Frame(parent, bg=C["bg"])
        tf.pack(fill="both", expand=True, padx=8, pady=(0,4))

        style = ttk.Style()
        style.configure("LR.Treeview",
                        background=C["surface"], foreground=C["text"],
                        fieldbackground=C["surface"], rowheight=24,
                        font=FMONO10)
        style.configure("LR.Treeview.Heading",
                        background="#4A1000", foreground="#FF8844",
                        font=("Arial",9,"bold"), relief="flat")
        style.map("LR.Treeview",
                  background=[("selected", "#CC0000")],
                  foreground=[("selected","white")])

        self.lr_tree = ttk.Treeview(tf, columns=cols, show="headings",
                                     style="LR.Treeview", selectmode="browse")
        widths = [150,110,110,100,60,130,160,120,180]
        for col, w in zip(cols, widths):
            self.lr_tree.heading(col, text=col)
            self.lr_tree.column(col, width=w, minwidth=50)

        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.lr_tree.yview)
        self.lr_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.lr_tree.pack(fill="both", expand=True)

        self.lr_count_lbl = tk.Label(parent, text="", font=FSML,
                                      fg=C["dim"], bg=C["bg"])
        self.lr_count_lbl.pack(anchor="e", padx=14, pady=(0,4))

    def _load_lr_log(self):
        q = self._lrfvar.get() if hasattr(self, "_lrfvar") else ""
        rows = self.db.get_liberty_risk_log(text_filter=q)
        for r in self.lr_tree.get_children():
            self.lr_tree.delete(r)
        for row in rows:
            self.lr_tree.insert("", "end", values=(
                row["timestamp"],
                row["last_name"],
                row["first_name"],
                row["edipi"],
                row["direction"],
                row["destination"],
                row["acknowledger_name"],
                row["acknowledger_edipi"],
                row["liberty_risk_reason"],
            ))
        total = self.db.liberty_risk_log_count()
        self.lr_count_lbl.config(
            text=f"{len(rows)} shown / {total} total acknowledgements")

    def _export_lr_log(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV","*.csv"),("All","*.*")],
            initialfile=f"liberty_risk_log_{datetime.date.today()}.csv",
            parent=self
        )
        if not path:
            return
        rows = self.db.get_liberty_risk_log()
        cols = ["timestamp","last_name","first_name","edipi","direction",
                "destination","acknowledger_name","acknowledger_edipi",
                "liberty_risk_reason"]
        with open(path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)
        messagebox.showinfo("Exported", f"Liberty Risk log saved:\n{path}",
                            parent=self)

    # ── Import / export ─────────────────────────────────────────────────

    def _import_csv(self):
        path = filedialog.askopenfilename(
            filetypes=[
                ("All supported formats", "*.xlsx *.xlsm *.csv *.tsv"),
                ("Excel workbooks",       "*.xlsx *.xlsm"),
                ("CSV / TSV files",       "*.csv *.tsv"),
                ("All files",             "*.*"),
            ],
            title="Select Roster File (JPERSTAT XLSX, CSV, or TSV)",
            parent=self
        )
        if not path:
            return
        try:
            imported, skipped, fmt = self.db.import_file(path)
            fmt_label = {
                "jperstat_xlsx": "JPERSTAT XLSX (PERSTAT sheet)",
                "jperstat_csv":  "JPERSTAT CSV",
                "legacy_csv":    "Legacy roster CSV",
            }.get(fmt, fmt)
            messagebox.showinfo(
                "Import Complete",
                f"Format detected: {fmt_label}\n\n"
                f"Imported:  {imported}\n"
                f"Skipped:   {skipped}",
                parent=self
            )
            self._after_edit()
        except Exception as e:
            messagebox.showerror("Import Error", str(e), parent=self)

    def _dl_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files","*.csv")],
            initialfile="cac_roster_template.csv", parent=self
        )
        if not path:
            return
        with open(path, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["SUBJECT_LAST_NAME","SUBJECT_FIRST_NAME",
                        "DATE_OF_BIRTH","PLACE OF BIRTH","EDIPI"])
            w.writerow(["WALKER","DEVAUGHN",
                        "10/19/1993","KINGSTON,JAM","1517290137"])
            w.writerow(["SMITH","JOHN",
                        "03/15/1988","NORFOLK,VA","9876543210"])
        messagebox.showinfo("Template Saved", f"Saved to:\n{path}", parent=self)

    def _export_audit(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files","*.csv"),("All","*.*")],
            initialfile=f"roster_audit_{datetime.date.today()}.csv",
            parent=self
        )
        if not path:
            return
        rows = self.db.get_audit_log()
        cols = ["timestamp","action","last_name","first_name","edipi",
                "fascn_key","changed_fields","old_values","new_values",
                "performed_by"]
        with open(path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)
        messagebox.showinfo("Exported", f"Audit log saved to:\n{path}",
                            parent=self)

    def _sort_roster(self, col):
        data = [(self.r_tree.set(c, col), c)
                for c in self.r_tree.get_children("")]
        self._r_sort_rev = (not self._r_sort_rev
                            if self._r_sort_col == col else False)
        self._r_sort_col = col
        data.sort(reverse=self._r_sort_rev)
        for i, (_, c) in enumerate(data):
            self.r_tree.move(c, "", i)




# =============================================================================
# DESTINATION MODAL  (shown on every CHECK OUT)
# =============================================================================

class DestinationModal(tk.Toplevel):
    """
    Shown immediately after a CHECK OUT is logged.
    Operator selects one of four tile options; the destination string is
    passed back via callback and written to the access_log row.

    Tiles highlight on hover and flash on selection.
    'Other' reveals a free-text field.
    """

    DESTINATIONS = [
        ("Official Business",  "💼", "#0057B8"),
        ("Deacon / Caritas",   "⛪", "#5B4A8A"),
        ("Restaurant / Mall",  "🍽", "#2A7A3B"),
        ("Other",              "✏️", "#8A6000"),
        ("R&R Leave",          "🏖", "#D4A017"),
        ("96-Hour Liberty",    "🏖", "#C77B30"),
    ]

    def __init__(self, parent, person: dict, callback):
        super().__init__(parent)
        self.title("Destination")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self.callback     = callback
        self._selected    = None
        self._flash_after = None

        self._build(person)
        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.after(50, self._center)

    def _center(self):
        self.update_idletasks()
        pw = self.master.winfo_rootx()
        py = self.master.winfo_rooty()
        w  = self.winfo_width()
        h  = self.winfo_height()
        self.geometry(
            f"+{pw + self.master.winfo_width()//2 - w//2}"
            f"+{py + self.master.winfo_height()//2 - h//2}"
        )

    def _build(self, person):
        tk.Frame(self, bg=C["red"], height=4).pack(fill="x")

        hdr = tk.Frame(self, bg=C["surface"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="▼  CHECK OUT", font=("Arial", 13, "bold"),
                 fg=C["red"], bg=C["surface"]).pack(side="left", padx=16)
        name = (f"{person.get('first_name','')} "
                f"{person.get('last_name','')}").strip()
        tk.Label(hdr, text=name, font=("Courier New", 11, "bold"),
                 fg=C["text"], bg=C["surface"]).pack(side="left", padx=4)

        tk.Label(self, text="Where are you headed?",
                 font=("Arial", 11, "bold"), fg=C["text"], bg=C["bg"]
                 ).pack(pady=(14, 6))

        tile_frame = tk.Frame(self, bg=C["bg"])
        tile_frame.pack(padx=20, pady=(0, 4))

        self._tiles       = {}
        self._other_var   = tk.StringVar()
        self._other_frame = tk.Frame(self, bg=C["bg"])

        for i, (label, icon, accent) in enumerate(self.DESTINATIONS):
            col = i % 2
            row = i // 2
            tile = tk.Frame(tile_frame, bg=C["surface2"],
                            highlightbackground=C["border"],
                            highlightthickness=2,
                            cursor="hand2", width=200, height=90)
            tile.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
            tile.pack_propagate(False)
            tile_frame.grid_columnconfigure(col, weight=1)

            inner = tk.Frame(tile, bg=C["surface2"])
            inner.place(relx=0.5, rely=0.5, anchor="center")
            tk.Label(inner, text=icon,  font=("Arial", 22),
                     bg=C["surface2"], fg=accent).pack()
            tk.Label(inner, text=label, font=("Arial", 10, "bold"),
                     bg=C["surface2"], fg=C["text"]).pack()

            self._tiles[label] = {"frame": tile, "accent": accent}

            def _enter(e, t=tile, a=accent, lbl=label):
                if self._selected != lbl:
                    t.config(highlightbackground=a, highlightthickness=3)
                    self._recolor(t, C["surface"], C["text"])

            def _leave(e, t=tile, lbl=label):
                if self._selected == lbl:
                    return
                t.config(highlightbackground=C["border"], highlightthickness=2)
                self._recolor(t, C["surface2"], C["text"])

            def _click(e, lbl=label, a=accent):
                self._select(lbl, a)

            for w in [tile, inner] + list(inner.winfo_children()):
                w.bind("<Enter>",    _enter)
                w.bind("<Leave>",    _leave)
                w.bind("<Button-1>", _click)

        # Other text field (hidden until Other is selected)
        tk.Label(self._other_frame, text="Specify destination:",
                 font=FSML, fg=C["dim"], bg=C["bg"]
                 ).pack(anchor="w", padx=4)
        self._other_entry = tk.Entry(
            self._other_frame, textvariable=self._other_var,
            font=("Courier New", 11),
            bg=C["surface2"], fg=C["bright"],
            insertbackground=C["bright"], relief="flat",
            highlightbackground=C["border"], highlightthickness=1,
            width=36)
        self._other_entry.pack(fill="x", padx=4, pady=(2, 6))
        self._other_entry.bind("<Return>", lambda _: self._confirm())

        # Buttons
        btn_row = tk.Frame(self, bg=C["surface"], pady=10)
        btn_row.pack(fill="x", side="bottom")
        tk.Button(btn_row, text="Skip", font=FSML,
                  bg=C["surface2"], fg=C["dim"], relief="flat",
                  padx=14, pady=5, cursor="hand2",
                  command=self._skip).pack(side="right", padx=8)
        self._confirm_btn = tk.Button(
            btn_row, text="Confirm  ▶",
            font=("Arial", 10, "bold"),
            bg=C["accent"], fg="white",
            activebackground=C["accent_light"],
            relief="flat", padx=16, pady=5, cursor="hand2",
            command=self._confirm, state="disabled")
        self._confirm_btn.pack(side="right", padx=4)

    def _recolor(self, widget, bg, fg):
        try:    widget.config(bg=bg)
        except tk.TclError: pass
        try:    widget.config(fg=fg)
        except tk.TclError: pass
        for child in widget.winfo_children():
            self._recolor(child, bg, fg)

    def _select(self, label: str, accent: str):
        # Reset all tiles
        for lbl, info in self._tiles.items():
            if lbl != label:
                t = info["frame"]
                t.config(highlightbackground=C["border"], highlightthickness=2)
                self._recolor(t, C["surface2"], C["text"])

        self._selected = label
        self._flash_tile(self._tiles[label]["frame"], accent)

        if label == "Other":
            self._other_frame.pack(padx=20, pady=(0, 6), fill="x",
                                   before=self.winfo_children()[-1])
            self._other_entry.focus_set()
        else:
            self._other_frame.pack_forget()
            self._other_var.set("")

        self._confirm_btn.config(state="normal")
        self.after(60, self._center)

    def _flash_tile(self, tile: tk.Frame, accent: str):
        """Rapidly alternate bg 4 times then settle on selected state."""
        if self._flash_after:
            self.after_cancel(self._flash_after)
        sequence = [accent, C["surface2"]] * 4 + [accent]

        def _step(idx):
            if idx >= len(sequence):
                tile.config(highlightbackground=accent, highlightthickness=3)
                self._recolor(tile, accent, C["bg"])
                return
            bg = sequence[idx]
            fg = C["bg"] if bg == accent else C["text"]
            tile.config(highlightbackground=bg, highlightthickness=3)
            self._recolor(tile, bg, fg)
            self._flash_after = self.after(55, lambda: _step(idx + 1))

        _step(0)

    def _confirm(self):
        if not self._selected:
            return
        if self._selected == "Other":
            dest = self._other_var.get().strip()
            if not dest:
                self._other_entry.config(highlightbackground=C["red"])
                self._other_entry.focus_set()
                return
            dest = f"Other: {dest}"
        else:
            dest = self._selected
        self.destroy()
        self.callback(dest)

    def _skip(self):
        self.destroy()
        self.callback("")


        self.callback("")


# =============================================================================
# LIBERTY RISK MODAL
# =============================================================================

class LibertyRiskModal(tk.Toplevel):
    """
    Shown whenever a Liberty Risk-flagged person scans their CAC.

    Phase 1 — RED ALERT: flashing warning screen with person's name and reason.
    Operator must click "I Acknowledge" to proceed.

    Phase 2 — SIGNATURE: operator enters their full name and EDIPI as
    digital acknowledgement. Clicking Confirm logs the event and
    fires the callback to proceed with check-in/out.

    The modal cannot be dismissed by clicking X — the operator must
    either sign or explicitly cancel, which also gets logged.
    """

    FLASH_COLORS = ["#CC0000", "#FF2222", "#990000", "#FF4444"]

    def __init__(self, parent, person: dict, direction: str,
                 destination: str, db, callback):
        super().__init__(parent)
        self.title("⚠  LIBERTY RISK ALERT")
        self.configure(bg="#1A0000")
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)

        self.person      = person
        self.direction   = direction
        self.destination = destination
        self.db          = db
        self.callback    = callback   # callback(acknowledged: bool)

        self._flash_idx   = 0
        self._flash_after = None
        self._phase       = 1         # 1=alert, 2=signature

        self._build_phase1()
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.after(50, self._center)
        self._start_flash()

    # ── Layout helpers ────────────────────────────────────────────────────────

    def _center(self):
        self.update_idletasks()
        pw = self.master.winfo_rootx()
        py = self.master.winfo_rooty()
        w  = self.winfo_width()
        h  = self.winfo_height()
        self.geometry(
            f"+{pw + self.master.winfo_width()//2 - w//2}"
            f"+{py + self.master.winfo_height()//2 - h//2}"
        )

    def _clear(self):
        for w in self.winfo_children():
            w.destroy()

    # ── Phase 1: Flashing alert ───────────────────────────────────────────────

    def _build_phase1(self):
        name = (f"{self.person.get('rank','')} "
                f"{self.person.get('first_name','')} "
                f"{self.person.get('last_name','')}").strip()
        reason = self.person.get("liberty_risk_reason", "") or "No reason specified"
        edipi  = self.person.get("edipi", "")
        dir_lbl = "CHECKING OUT" if self.direction == "OUT" else "CHECKING IN"

        self._alert_frame = tk.Frame(self, bg="#CC0000")
        self._alert_frame.pack(fill="both", expand=True)

        # Warning icon row
        tk.Label(self._alert_frame,
                 text="⚠",
                 font=("Arial", 56),
                 fg="#FFFF00", bg="#CC0000"
                 ).pack(pady=(24, 0))

        tk.Label(self._alert_frame,
                 text="LIBERTY RISK",
                 font=("Arial", 28, "bold"),
                 fg="#FFFFFF", bg="#CC0000"
                 ).pack()

        tk.Frame(self._alert_frame, bg="#FF6666", height=2).pack(fill="x",
                 padx=24, pady=8)

        # Person info
        tk.Label(self._alert_frame,
                 text=name,
                 font=("Courier New", 18, "bold"),
                 fg="#FFFFFF", bg="#CC0000"
                 ).pack()

        tk.Label(self._alert_frame,
                 text=f"DoD ID: {edipi}",
                 font=("Courier New", 12),
                 fg="#FFCCCC", bg="#CC0000"
                 ).pack()

        tk.Label(self._alert_frame,
                 text=f"— {dir_lbl} —",
                 font=("Arial", 11, "bold"),
                 fg="#FFFF99", bg="#CC0000"
                 ).pack(pady=(10, 0))

        tk.Frame(self._alert_frame, bg="#FF6666", height=2).pack(fill="x",
                 padx=24, pady=8)

        # Reason box
        reason_frame = tk.Frame(self._alert_frame, bg="#990000",
                                padx=16, pady=10)
        reason_frame.pack(fill="x", padx=20, pady=(0, 8))
        tk.Label(reason_frame,
                 text="REASON FOR DESIGNATION:",
                 font=("Arial", 9, "bold"),
                 fg="#FFAAAA", bg="#990000"
                 ).pack(anchor="w")
        tk.Label(reason_frame,
                 text=reason,
                 font=("Arial", 11),
                 fg="#FFFFFF", bg="#990000",
                 wraplength=380, justify="left"
                 ).pack(anchor="w", pady=(4, 0))

        # Instruction
        tk.Label(self._alert_frame,
                 text="This event requires Authorizing Official acknowledgement.\n"
                      "Your name and DoD ID will be recorded as your signature.",
                 font=("Arial", 9),
                 fg="#FFCCCC", bg="#CC0000",
                 justify="center"
                 ).pack(pady=(4, 12))

        # Buttons
        btn_row = tk.Frame(self._alert_frame, bg="#CC0000")
        btn_row.pack(pady=(0, 20))

        tk.Button(btn_row,
                  text="Cancel",
                  font=("Arial", 10),
                  bg="#660000", fg="#FFAAAA",
                  activebackground="#440000",
                  relief="flat", padx=16, pady=7,
                  cursor="hand2",
                  command=self._cancel
                  ).pack(side="left", padx=8)

        tk.Button(btn_row,
                  text="I Acknowledge — Continue  ▶",
                  font=("Arial", 11, "bold"),
                  bg="#FFFF00", fg="#330000",
                  activebackground="#FFEE00",
                  relief="flat", padx=20, pady=7,
                  cursor="hand2",
                  command=self._phase2
                  ).pack(side="left", padx=8)

    def _start_flash(self):
        """Cycle the alert frame background through red shades."""
        if self._phase != 1:
            return
        color = self.FLASH_COLORS[self._flash_idx % len(self.FLASH_COLORS)]
        self._flash_idx += 1
        try:
            self._alert_frame.config(bg=color)
            # Also flash child labels that have the old bg
            for w in self._alert_frame.winfo_children():
                try:
                    if w.cget("bg") in self.FLASH_COLORS or \
                       w.cget("bg") == "#CC0000":
                        w.config(bg=color)
                except tk.TclError:
                    pass
        except tk.TclError:
            return
        self._flash_after = self.after(400, self._start_flash)

    # ── Phase 2: Signature ────────────────────────────────────────────────────

    def _phase2(self):
        """Stop flashing, switch to signature entry."""
        self._phase = 2
        if self._flash_after:
            self.after_cancel(self._flash_after)
            self._flash_after = None

        self._clear()
        self.configure(bg=C["bg"])

        # Header
        tk.Frame(self, bg="#CC0000", height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=C["surface"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr,
                 text="⚠  LIBERTY RISK — AUTHORIZING OFFICIAL ACKNOWLEDGEMENT",
                 font=("Arial", 11, "bold"),
                 fg="#FF4444", bg=C["surface"]
                 ).pack(side="left", padx=16)

        # Reminder of who is being checked out
        name = (f"{self.person.get('rank','')} "
                f"{self.person.get('first_name','')} "
                f"{self.person.get('last_name','')}").strip()
        info = tk.Frame(self, bg=C["surface2"], padx=16, pady=10)
        info.pack(fill="x", padx=14, pady=(12, 0))
        tk.Label(info,
                 text=f"Personnel: {name}  |  "
                      f"DoD ID: {self.person.get('edipi','')}  |  "
                      f"Action: {self.direction}",
                 font=FMONO10, fg=C["gold"], bg=C["surface2"]
                 ).pack(anchor="w")
        if self.destination:
            tk.Label(info,
                     text=f"Destination: {self.destination}",
                     font=FMONO10, fg=C["dim"], bg=C["surface2"]
                     ).pack(anchor="w")

        # Instruction
        body = tk.Frame(self, bg=C["bg"], padx=20, pady=14)
        body.pack(fill="both")

        tk.Label(body,
                 text="By entering your information below, you acknowledge that you are\n"
                      "aware this individual is designated as a Liberty Risk and accept\n"
                      "responsibility for authorizing this access event.",
                 font=("Arial", 10),
                 fg=C["text"], bg=C["bg"],
                 justify="left"
                 ).pack(anchor="w", pady=(0, 14))

        # Signature fields
        self._sig_name_var  = tk.StringVar()
        self._sig_edipi_var = tk.StringVar()

        for label, var, width in [
            ("Your Full Name *",   self._sig_name_var,  30),
            ("Your DoD ID (EDIPI) *", self._sig_edipi_var, 18),
        ]:
            row = tk.Frame(body, bg=C["bg"])
            row.pack(fill="x", pady=4)
            tk.Label(row, text=label, font=FBOLD,
                     fg=C["text"], bg=C["bg"], width=22, anchor="w"
                     ).pack(side="left")
            e = tk.Entry(row, textvariable=var,
                         font=FID,
                         bg=C["surface2"], fg=C["bright"],
                         insertbackground=C["bright"],
                         relief="flat",
                         highlightbackground=C["border"],
                         highlightthickness=1,
                         width=width)
            e.pack(side="left", padx=(8, 0))

        self._sig_err = tk.Label(body, text="",
                                  font=("Arial", 9),
                                  fg=C["red"], bg=C["bg"])
        self._sig_err.pack(anchor="w", pady=(4, 0))

        tk.Label(body,
                 text="* Required — this constitutes your digital signature",
                 font=("Arial", 8), fg=C["dim"], bg=C["bg"]
                 ).pack(anchor="w", pady=(2, 0))

        # Buttons
        btn_row = tk.Frame(self, bg=C["surface"], pady=10)
        btn_row.pack(fill="x", side="bottom")
        tk.Button(btn_row,
                  text="Cancel",
                  font=FSML,
                  bg=C["surface2"], fg=C["dim"],
                  relief="flat", padx=14, pady=6,
                  cursor="hand2",
                  command=self._cancel
                  ).pack(side="right", padx=8)
        tk.Button(btn_row,
                  text="Confirm & Log  ✔",
                  font=("Arial", 10, "bold"),
                  bg="#CC0000", fg="white",
                  activebackground="#AA0000",
                  relief="flat", padx=16, pady=6,
                  cursor="hand2",
                  command=self._confirm
                  ).pack(side="right", padx=4)

        self.after(60, self._center)

    def _confirm(self):
        sig_name  = self._sig_name_var.get().strip()
        sig_edipi = re.sub(r"\D", "", self._sig_edipi_var.get())

        if not sig_name:
            self._sig_err.config(text="⚠  Full name is required.")
            return
        if not sig_edipi:
            self._sig_err.config(text="⚠  DoD ID (EDIPI) is required.")
            return
        if len(sig_edipi) < 7:
            self._sig_err.config(text="⚠  EDIPI must be at least 7 digits.")
            return

        # Log the acknowledgement
        self.db.log_liberty_risk_acknowledgement(
            fascn_key        = self.person.get("fascn_key", ""),
            person           = self.person,
            direction        = self.direction,
            destination      = self.destination,
            acknowledger_name  = sig_name,
            acknowledger_edipi = sig_edipi,
        )
        self.destroy()
        self.callback(True, sig_name, sig_edipi)

    def _cancel(self):
        if self._flash_after:
            self.after_cancel(self._flash_after)
        self.destroy()
        self.callback(False, "", "")


# =============================================================================
# CALENDAR PICKER  (stdlib `calendar` module — no third-party deps)
# =============================================================================

class CalendarPicker(tk.Frame):
    """
    Self-contained month-grid calendar. Click a day to select it; the prev/next
    arrows navigate months. Calls on_select(datetime.date) on each pick.
    Built entirely from the stdlib `calendar` module so the portable bundle
    needs no extra packages.
    """
    DOW = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]
    MONTHS = ["January", "February", "March", "April", "May", "June", "July",
              "August", "September", "October", "November", "December"]

    def __init__(self, parent, initial: datetime.date = None,
                 on_select=None, min_date: datetime.date = None):
        super().__init__(parent, bg=C["surface"])
        self.on_select = on_select
        self.min_date  = min_date
        today = datetime.date.today()
        self._view = (initial or today).replace(day=1)
        self._selected = initial
        self._build()
        self._render()

    def _build(self):
        hdr = tk.Frame(self, bg=C["surface2"])
        hdr.pack(fill="x")
        tk.Button(hdr, text="◀", font=FBOLD, bg=C["surface2"], fg=C["gold"],
                  relief="flat", cursor="hand2", bd=0, padx=10,
                  activebackground=C["border"], command=self._prev_month
                  ).pack(side="left")
        self._title = tk.Label(hdr, text="", font=FBOLD,
                               fg=C["text"], bg=C["surface2"])
        self._title.pack(side="left", expand=True)
        tk.Button(hdr, text="▶", font=FBOLD, bg=C["surface2"], fg=C["gold"],
                  relief="flat", cursor="hand2", bd=0, padx=10,
                  activebackground=C["border"], command=self._next_month
                  ).pack(side="right")

        dow = tk.Frame(self, bg=C["surface"])
        dow.pack(fill="x", pady=(4, 0))
        for i, d in enumerate(self.DOW):
            tk.Label(dow, text=d, font=("Arial", 8, "bold"),
                     fg=C["dim"], bg=C["surface"], width=4
                     ).grid(row=0, column=i, padx=1, pady=1)

        self._grid = tk.Frame(self, bg=C["surface"])
        self._grid.pack(fill="both", pady=(0, 4))
        self._day_btns = []
        for r in range(6):
            row = []
            for c in range(7):
                b = tk.Button(self._grid, text="", font=FSML, width=4, height=1,
                              relief="flat", bd=0, bg=C["surface"], fg=C["text"],
                              activebackground=C["accent"], cursor="hand2")
                b.grid(row=r, column=c, padx=1, pady=1)
                row.append(b)
            self._day_btns.append(row)

    def _render(self):
        y, m = self._view.year, self._view.month
        self._title.config(text=f"{self.MONTHS[m - 1]} {y}")
        cal = calendar.Calendar(firstweekday=0)
        weeks = cal.monthdayscalendar(y, m)
        today = datetime.date.today()
        for r in range(6):
            for c in range(7):
                b = self._day_btns[r][c]
                if r < len(weeks) and weeks[r][c] != 0:
                    day = weeks[r][c]
                    d = datetime.date(y, m, day)
                    disabled = (self.min_date is not None and d < self.min_date)
                    b.config(
                        text=str(day),
                        state="disabled" if disabled else "normal",
                        bg=self._day_bg(d, today),
                        fg=self._day_fg(d, today, disabled),
                        command=(lambda dd=d: self._pick(dd)) if not disabled else 0
                    )
                else:
                    b.config(text="", state="disabled",
                             bg=C["surface"], fg=C["surface"], command=0)

    def _day_bg(self, d, today):
        if self._selected and d == self._selected:
            return C["accent"]
        if d == today:
            return C["surface2"]
        return C["surface"]

    def _day_fg(self, d, today, disabled):
        if disabled:
            return C["border"]
        if self._selected and d == self._selected:
            return C["bright"]
        if d == today:
            return C["gold"]
        return C["text"]

    def _pick(self, d):
        self._selected = d
        self._render()
        if self.on_select:
            self.on_select(d)

    def _prev_month(self):
        y, m = self._view.year, self._view.month
        m -= 1
        if m < 1:
            m = 12; y -= 1
        self._view = datetime.date(y, m, 1)
        self._render()

    def _next_month(self):
        y, m = self._view.year, self._view.month
        m += 1
        if m > 12:
            m = 1; y += 1
        self._view = datetime.date(y, m, 1)
        self._render()

    def get(self):
        return self._selected

    def set(self, d):
        self._selected = d
        self._view = d.replace(day=1)
        self._render()


# =============================================================================
# LEAVE DATE-RANGE MODAL
# =============================================================================

class LeaveDateRangeModal(tk.Toplevel):
    """
    Shown after the operator chooses R&R Leave or 96-Hour Liberty on a CHECK OUT.
    Two calendars (start + end). On confirm -> callback(start_iso, end_iso).
    Cancel -> callback(None, None).
    """
    def __init__(self, parent, person, leave_type, callback):
        super().__init__(parent)
        self.title(f"{leave_type} — Select Dates")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self.person     = person
        self.leave_type = leave_type
        self.callback   = callback
        self._start = datetime.date.today()
        self._end   = datetime.date.today()
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.after(50, self._center)

    def _center(self):
        self.update_idletasks()
        pw = self.master.winfo_rootx(); py = self.master.winfo_rooty()
        w = self.winfo_width(); h = self.winfo_height()
        self.geometry(f"+{pw + self.master.winfo_width()//2 - w//2}"
                      f"+{py + self.master.winfo_height()//2 - h//2}")

    def _build(self):
        accent = C["gold"] if self.leave_type == LEAVE_RR else "#8A6000"
        tk.Frame(self, bg=accent, height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=C["surface"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"🏖  {self.leave_type}",
                 font=("Arial", 13, "bold"), fg=accent, bg=C["surface"]
                 ).pack(side="left", padx=16)
        name = (f"{self.person.get('first_name','')} "
                f"{self.person.get('last_name','')}").strip()
        tk.Label(hdr, text=name, font=("Courier New", 11, "bold"),
                 fg=C["text"], bg=C["surface"]).pack(side="left", padx=4)

        tk.Label(self, text="Select the start and end dates of this leave period.",
                 font=FSML, fg=C["dim"], bg=C["bg"]).pack(pady=(10, 4))

        cals = tk.Frame(self, bg=C["bg"])
        cals.pack(padx=16, pady=4)

        start_col = tk.Frame(cals, bg=C["bg"])
        start_col.pack(side="left", padx=8)
        tk.Label(start_col, text="FROM (start date)", font=FBOLD,
                 fg=C["green"], bg=C["bg"]).pack(pady=(0, 4))
        self._start_lbl = tk.Label(start_col, text=self._start.isoformat(),
                 font=FMONO, fg=C["text"], bg=C["surface2"], pady=4)
        self._start_lbl.pack(fill="x", pady=(0, 4))
        self._start_cal = CalendarPicker(start_col, initial=self._start,
                 on_select=self._on_start, min_date=datetime.date.today())
        self._start_cal.pack()

        end_col = tk.Frame(cals, bg=C["bg"])
        end_col.pack(side="left", padx=8)
        tk.Label(end_col, text="TO (end date)", font=FBOLD,
                 fg=C["red"], bg=C["bg"]).pack(pady=(0, 4))
        self._end_lbl = tk.Label(end_col, text=self._end.isoformat(),
                 font=FMONO, fg=C["text"], bg=C["surface2"], pady=4)
        self._end_lbl.pack(fill="x", pady=(0, 4))
        self._end_cal = CalendarPicker(end_col, initial=self._end,
                 on_select=self._on_end, min_date=datetime.date.today())
        self._end_cal.pack()

        self._dur_lbl = tk.Label(self, text="", font=FSML, fg=C["gold"], bg=C["bg"])
        self._dur_lbl.pack(pady=(6, 0))
        self._update_duration()

        btn_row = tk.Frame(self, bg=C["surface"], pady=10)
        btn_row.pack(fill="x", side="bottom")
        tk.Button(btn_row, text="Cancel", font=FSML, bg=C["surface2"], fg=C["dim"],
                  relief="flat", padx=16, pady=6, cursor="hand2",
                  command=self._cancel).pack(side="right", padx=8)
        tk.Button(btn_row, text="Confirm Leave Dates  ▶",
                  font=("Arial", 10, "bold"), bg=C["accent"], fg="white",
                  activebackground=C["accent_light"], relief="flat",
                  padx=16, pady=6, cursor="hand2",
                  command=self._confirm).pack(side="right", padx=4)

    def _on_start(self, d):
        self._start = d
        self._start_lbl.config(text=d.isoformat())
        if self._end < self._start:
            self._end = self._start
            self._end_cal.set(self._end)
            self._end_lbl.config(text=self._end.isoformat())
        self._update_duration()

    def _on_end(self, d):
        self._end = d
        self._end_lbl.config(text=d.isoformat())
        self._update_duration()

    def _update_duration(self):
        if self._end < self._start:
            self._dur_lbl.config(text="⚠  End date is before start date.",
                                 fg=C["red"])
            return
        days = (self._end - self._start).days + 1
        self._dur_lbl.config(
            text=f"Duration: {days} day{'s' if days != 1 else ''}  "
                 f"({self._start.isoformat()} → {self._end.isoformat()})",
            fg=C["gold"])

    def _confirm(self):
        if self._end < self._start:
            messagebox.showwarning("Invalid Dates",
                "The end date cannot be before the start date.", parent=self)
            return
        self.destroy()
        self.callback(self._start.isoformat(), self._end.isoformat())

    def _cancel(self):
        self.destroy()
        self.callback(None, None)


# =============================================================================
# LEAVE RETURN CONFIRMATION MODAL
# =============================================================================

class LeaveReturnModal(tk.Toplevel):
    """
    Shown when a person with an ACTIVE leave record scans/checks back in.
    "You are about to be checked back in from a leave status, is this correct?"
    Shows the original leave dates and Yes/No buttons. callback(confirmed: bool).
    """
    def __init__(self, parent, person, leave_record, callback):
        super().__init__(parent)
        self.title("Confirm Return From Leave")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self.person = person
        self.rec = leave_record
        self.callback = callback
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._no)
        self.after(50, self._center)

    def _center(self):
        self.update_idletasks()
        pw = self.master.winfo_rootx(); py = self.master.winfo_rooty()
        w = self.winfo_width(); h = self.winfo_height()
        self.geometry(f"+{pw + self.master.winfo_width()//2 - w//2}"
                      f"+{py + self.master.winfo_height()//2 - h//2}")

    def _build(self):
        tk.Frame(self, bg=C["green"], height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=C["surface"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="↩  RETURN FROM LEAVE",
                 font=("Arial", 13, "bold"), fg=C["green"], bg=C["surface"]
                 ).pack(side="left", padx=16)

        body = tk.Frame(self, bg=C["bg"], padx=20, pady=16)
        body.pack(fill="both")
        name = (f"{self.person.get('rank','')} "
                f"{self.person.get('first_name','')} "
                f"{self.person.get('last_name','')}").strip()
        tk.Label(body, text=name, font=("Courier New", 14, "bold"),
                 fg=C["text"], bg=C["bg"]).pack(anchor="w")
        tk.Label(body, text=f"DoD ID: {self.person.get('edipi','')}",
                 font=FMONO10, fg=C["dim"], bg=C["bg"]
                 ).pack(anchor="w", pady=(0, 12))

        tk.Label(body,
                 text="You are about to be checked back in from a leave status,\n"
                      "is this correct?",
                 font=("Arial", 12, "bold"), fg=C["bright"], bg=C["bg"],
                 justify="left").pack(anchor="w", pady=(0, 12))

        panel = tk.Frame(body, bg=C["surface2"],
                         highlightbackground=C["border"], highlightthickness=1)
        panel.pack(fill="x", pady=(0, 8))
        lt = self.rec.get("leave_type", "")
        start = self.rec.get("start_date", "")
        end = self.rec.get("end_date", "")
        dep = self.rec.get("departure_ts", "")
        try:
            d0 = datetime.date.fromisoformat(start)
            d1 = datetime.date.fromisoformat(end)
            days = (d1 - d0).days + 1
        except Exception:
            days = "?"
        accent = C["gold"] if lt == LEAVE_RR else "#CC9933"
        tk.Label(panel, text=f"  {lt}", font=FBOLD, fg=accent,
                 bg=C["surface2"]).pack(anchor="w", padx=10, pady=(8, 2))
        for lbl, val in [("Leave start:", start), ("Leave end:", end),
                         ("Duration:", f"{days} day{'s' if days != 1 else ''}"),
                         ("Departed:", dep.replace("T", " ") if dep else "—")]:
            row = tk.Frame(panel, bg=C["surface2"])
            row.pack(fill="x", padx=10, pady=1)
            tk.Label(row, text=lbl, font=FSML, fg=C["dim"], bg=C["surface2"],
                     width=12, anchor="w").pack(side="left")
            tk.Label(row, text=val, font=FMONO10, fg=C["text"],
                     bg=C["surface2"], anchor="w").pack(side="left")
        tk.Frame(panel, bg=C["surface2"], height=6).pack()

        btn_row = tk.Frame(self, bg=C["surface"], pady=12)
        btn_row.pack(fill="x", side="bottom")
        tk.Button(btn_row, text="NO", font=("Arial", 11, "bold"),
                  bg=C["red"], fg="white", activebackground="#C02020",
                  relief="flat", padx=30, pady=8, cursor="hand2",
                  command=self._no).pack(side="right", padx=8)
        tk.Button(btn_row, text="YES — Check In", font=("Arial", 11, "bold"),
                  bg=C["green"], fg="white", activebackground="#00A368",
                  relief="flat", padx=30, pady=8, cursor="hand2",
                  command=self._yes).pack(side="right", padx=4)

    def _yes(self):
        self.destroy()
        self.callback(True)

    def _no(self):
        self.destroy()
        self.callback(False)


class LeaveEditModal(tk.Toplevel):
    """Admin edit of an existing leave record's start/end dates."""
    def __init__(self, parent, record, callback):
        super().__init__(parent)
        self.title("Edit Leave Dates")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self.record = record
        self.callback = callback
        try:
            self._start = datetime.date.fromisoformat(record.get("start_date", ""))
        except Exception:
            self._start = datetime.date.today()
        try:
            self._end = datetime.date.fromisoformat(record.get("end_date", ""))
        except Exception:
            self._end = self._start
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.after(50, self._center)

    def _center(self):
        self.update_idletasks()
        pw = self.master.winfo_rootx(); py = self.master.winfo_rooty()
        w = self.winfo_width(); h = self.winfo_height()
        self.geometry(f"+{pw + self.master.winfo_width()//2 - w//2}"
                      f"+{py + self.master.winfo_height()//2 - h//2}")

    def _build(self):
        tk.Frame(self, bg=C["gold"], height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=C["surface"], pady=10)
        hdr.pack(fill="x")
        name = (f"{self.record.get('first_name','')} "
                f"{self.record.get('last_name','')}").strip()
        tk.Label(hdr,
                 text=f"Edit Leave Dates  -  {self.record.get('leave_type','')}",
                 font=("Arial", 12, "bold"), fg=C["gold"], bg=C["surface"]
                 ).pack(side="left", padx=16)
        tk.Label(hdr, text=name, font=("Courier New", 11, "bold"),
                 fg=C["text"], bg=C["surface"]).pack(side="left", padx=4)

        cals = tk.Frame(self, bg=C["bg"])
        cals.pack(padx=16, pady=10)
        start_col = tk.Frame(cals, bg=C["bg"]); start_col.pack(side="left", padx=8)
        tk.Label(start_col, text="FROM (start date)", font=FBOLD,
                 fg=C["green"], bg=C["bg"]).pack(pady=(0, 4))
        self._start_lbl = tk.Label(start_col, text=self._start.isoformat(),
                 font=FMONO, fg=C["text"], bg=C["surface2"], pady=4)
        self._start_lbl.pack(fill="x", pady=(0, 4))
        self._start_cal = CalendarPicker(start_col, initial=self._start,
                 on_select=self._on_start)
        self._start_cal.pack()

        end_col = tk.Frame(cals, bg=C["bg"]); end_col.pack(side="left", padx=8)
        tk.Label(end_col, text="TO (end date)", font=FBOLD,
                 fg=C["red"], bg=C["bg"]).pack(pady=(0, 4))
        self._end_lbl = tk.Label(end_col, text=self._end.isoformat(),
                 font=FMONO, fg=C["text"], bg=C["surface2"], pady=4)
        self._end_lbl.pack(fill="x", pady=(0, 4))
        self._end_cal = CalendarPicker(end_col, initial=self._end,
                 on_select=self._on_end)
        self._end_cal.pack()

        self._dur_lbl = tk.Label(self, text="", font=FSML, fg=C["gold"], bg=C["bg"])
        self._dur_lbl.pack(pady=(6, 0))
        self._update_duration()

        btn_row = tk.Frame(self, bg=C["surface"], pady=10)
        btn_row.pack(fill="x", side="bottom")
        tk.Button(btn_row, text="Cancel", font=FSML, bg=C["surface2"], fg=C["dim"],
                  relief="flat", padx=16, pady=6, cursor="hand2",
                  command=self._cancel).pack(side="right", padx=8)
        tk.Button(btn_row, text="Save Dates", font=("Arial", 10, "bold"),
                  bg=C["accent"], fg="white", activebackground=C["accent_light"],
                  relief="flat", padx=16, pady=6, cursor="hand2",
                  command=self._save).pack(side="right", padx=4)

    def _on_start(self, d):
        self._start = d
        self._start_lbl.config(text=d.isoformat())
        if self._end < self._start:
            self._end = self._start
            self._end_cal.set(self._end)
            self._end_lbl.config(text=self._end.isoformat())
        self._update_duration()

    def _on_end(self, d):
        self._end = d
        self._end_lbl.config(text=d.isoformat())
        self._update_duration()

    def _update_duration(self):
        if self._end < self._start:
            self._dur_lbl.config(text="End date is before start date.", fg=C["red"])
            return
        days = (self._end - self._start).days + 1
        self._dur_lbl.config(
            text=f"Duration: {days} day{'s' if days != 1 else ''}  "
                 f"({self._start.isoformat()} to {self._end.isoformat()})",
            fg=C["gold"])

    def _save(self):
        if self._end < self._start:
            messagebox.showwarning("Invalid Dates",
                "The end date cannot be before the start date.", parent=self)
            return
        self.destroy()
        self.callback(self._start.isoformat(), self._end.isoformat())

    def _cancel(self):
        self.destroy()
        self.callback(None, None)


class RecordUpdateModal(tk.Toplevel):
    """Flashing-red prompt shown when a scanned record is missing critical
    fields (EDIPI, Last Name, First Name, Unit/Department). Lets the operator
    fill them in before the check-in/out proceeds."""
    FLASH = ["#CC0000", "#FF2222", "#990000", "#FF4444"]

    def __init__(self, parent, person, missing, db, callback):
        super().__init__(parent)
        self.title("Record Incomplete")
        self.configure(bg="#1A0000")
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self.person = person
        self.missing = missing
        self.db = db
        self.callback = callback
        self._flash_idx = 0
        self._flash_after = None
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.after(50, self._center)
        self._flash()

    def _center(self):
        self.update_idletasks()
        pw = self.master.winfo_rootx(); py = self.master.winfo_rooty()
        w = self.winfo_width(); h = self.winfo_height()
        self.geometry(f"+{pw + self.master.winfo_width()//2 - w//2}"
                      f"+{py + self.master.winfo_height()//2 - h//2}")

    def _build(self):
        self._bar = tk.Frame(self, bg="#FFFF00", height=4)
        self._bar.pack(fill="x")
        self._hdr = tk.Frame(self, bg="#CC0000", pady=12)
        self._hdr.pack(fill="x")
        tk.Label(self._hdr, text="RECORD INCOMPLETE",
                 font=("Arial", 18, "bold"), fg="#FFFFFF", bg="#CC0000").pack()
        name = (f"{self.person.get('rank','')} "
                f"{self.person.get('first_name','') or '(no first name)'} "
                f"{self.person.get('last_name','') or '(no last name)'}").strip()
        tk.Label(self._hdr, text=name, font=("Courier New", 12, "bold"),
                 fg="#FFEEEE", bg="#CC0000").pack()

        body = tk.Frame(self, bg=C["bg"], padx=20, pady=14)
        body.pack(fill="both")
        tk.Label(body,
                 text="This record is missing required information. "
                      "Please update it before continuing.",
                 font=("Arial", 10), fg=C["text"], bg=C["bg"],
                 wraplength=420, justify="left").pack(anchor="w", pady=(0, 4))
        tk.Label(body, text="Missing: " + ", ".join(self.missing),
                 font=("Arial", 9, "bold"), fg="#FF6666", bg=C["bg"]
                 ).pack(anchor="w", pady=(0, 10))

        self._vars = {}
        rows = [("DoD ID (EDIPI)", "edipi"), ("Last Name", "last_name"),
                ("First Name", "first_name"), ("Unit / Department", "unit")]
        for label, key in rows:
            row = tk.Frame(body, bg=C["bg"])
            row.pack(fill="x", pady=3)
            is_missing = label in self.missing
            tk.Label(row, text=label + ":", font=FBOLD,
                     fg=("#FF8888" if is_missing else C["dim"]), bg=C["bg"],
                     width=18, anchor="w").pack(side="left")
            var = tk.StringVar(value=str(self.person.get(key, "") or ""))
            self._vars[key] = var
            tk.Entry(row, textvariable=var, font=FID,
                     bg=C["surface2"], fg=C["bright"],
                     insertbackground=C["bright"], relief="flat",
                     highlightbackground=("#FF4444" if is_missing else C["border"]),
                     highlightthickness=2, width=26).pack(side="left", padx=(8, 0))

        self._err = tk.Label(body, text="", font=("Arial", 9),
                             fg=C["red"], bg=C["bg"])
        self._err.pack(anchor="w", pady=(4, 0))

        btn = tk.Frame(self, bg=C["surface"], pady=10)
        btn.pack(fill="x", side="bottom")
        tk.Button(btn, text="Skip for now", font=FSML,
                  bg=C["surface2"], fg=C["dim"], relief="flat",
                  padx=14, pady=6, cursor="hand2",
                  command=self._skip).pack(side="right", padx=8)
        tk.Button(btn, text="Save & Continue", font=("Arial", 10, "bold"),
                  bg=C["green"], fg="white", activebackground="#00A368",
                  relief="flat", padx=16, pady=6, cursor="hand2",
                  command=self._save).pack(side="right", padx=4)

    def _flash(self):
        col = self.FLASH[self._flash_idx % len(self.FLASH)]
        self._flash_idx += 1
        try:
            self._hdr.config(bg=col)
            for w in self._hdr.winfo_children():
                try:
                    w.config(bg=col)
                except tk.TclError:
                    pass
        except tk.TclError:
            return
        self._flash_after = self.after(400, self._flash)

    def _stop_flash(self):
        if self._flash_after:
            self.after_cancel(self._flash_after)
            self._flash_after = None

    def _save(self):
        edipi = self._vars["edipi"].get().strip()
        last = self._vars["last_name"].get().strip()
        first = self._vars["first_name"].get().strip()
        unit = self._vars["unit"].get().strip()
        key = self.person.get("fascn_key", "")
        data = dict(self.person)
        data["edipi"] = edipi
        data["last_name"] = last
        data["first_name"] = first
        try:
            self.db.update_person(key, data, performed_by="operator")
            self.db.update_unit(key, unit)
        except Exception as exc:
            self._err.config(text=f"Could not save: {exc}")
            return
        self._stop_flash()
        self.destroy()
        self.callback()

    def _skip(self):
        self._stop_flash()
        self.destroy()
        self.callback()


class AuthorizingOfficialSignModal(tk.Toplevel):
    def __init__(self, parent, heading, callback):
        super().__init__(parent)
        self.callback = callback
        self.title("Authorizing Official Approval")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        wrap = tk.Frame(self, bg=C["bg"], padx=18, pady=16)
        wrap.pack(fill="both", expand=True)
        tk.Label(wrap, text=heading, font=FBOLD, fg=C["gold"],
                 bg=C["bg"], wraplength=380, justify="left").pack(anchor="w")
        tk.Label(wrap,
                 text="A authorizing official must authorize this action. It is recorded "
                      "in the audit trail.",
                 font=FSML, fg=C["dim"], bg=C["bg"], wraplength=380,
                 justify="left").pack(anchor="w", pady=(2, 10))
        self.v_name = tk.StringVar()
        self.v_edipi = tk.StringVar()
        self.v_reason = tk.StringVar()
        rows = [("Authorizing Official Name", self.v_name),
                ("Authorizing Official DoD ID (EDIPI)", self.v_edipi),
                ("Reason", self.v_reason)]
        first = None
        for label, var in rows:
            row = tk.Frame(wrap, bg=C["bg"])
            row.pack(fill="x", pady=3)
            tk.Label(row, text=label, font=FSML, fg=C["text"], bg=C["bg"],
                     width=24, anchor="w").pack(side="left")
            e = tk.Entry(row, textvariable=var, font=FMONO10, bg=C["surface2"],
                         fg=C["text"], insertbackground=C["text"], relief="flat")
            e.pack(side="left", fill="x", expand=True, ipady=3)
            if first is None:
                first = e
        self.err = tk.Label(wrap, text="", font=FSML, fg=C["red"], bg=C["bg"])
        self.err.pack(anchor="w", pady=(6, 0))
        btns = tk.Frame(wrap, bg=C["bg"])
        btns.pack(fill="x", pady=(12, 0))
        tk.Button(btns, text="Authorize", command=self._ok, font=FBOLD,
                  bg=C["green"], fg="#06210B", relief="flat", padx=14, pady=6,
                  activebackground=C["green"], cursor="hand2").pack(side="left")
        tk.Button(btns, text="Cancel", command=self._cancel, font=FSML,
                  bg=C["surface2"], fg=C["text"], relief="flat", padx=12,
                  pady=6, cursor="hand2").pack(side="right")
        if first is not None:
            self.after(120, lambda: first.focus_set())
        self.bind("<Escape>", lambda e: self._cancel())

    def _ok(self):
        name = self.v_name.get().strip()
        edipi = self.v_edipi.get().strip()
        reason = self.v_reason.get().strip()
        if not name or not edipi or not reason:
            self.err.config(text="All fields are required.")
            return
        cb = self.callback
        self.destroy()
        cb(True, name, edipi, reason)

    def _cancel(self):
        cb = self.callback
        self.destroy()
        cb(False, "", "", "")


class GroupCheckOutModal(tk.Toplevel):
    def __init__(self, parent, seed_person, db, on_finalize):
        super().__init__(parent)
        self.app = parent
        self.db = db
        self.on_finalize = on_finalize
        self.members = []
        self.override_sig = None
        self.title("Group Check-Out")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        wrap = tk.Frame(self, bg=C["bg"], padx=18, pady=16)
        wrap.pack(fill="both", expand=True)
        tk.Label(wrap, text="GROUP CHECK-OUT", font=FTITLE,
                 fg=C["out_fg"], bg=C["bg"]).pack(anchor="w")
        tk.Label(wrap,
                 text="Scan each person departing together. A minimum of two is "
                      "required, or use Authorizing Official Override for an approved solo "
                      "departure.",
                 font=FSML, fg=C["dim"], bg=C["bg"], wraplength=440,
                 justify="left").pack(anchor="w", pady=(2, 10))
        scanrow = tk.Frame(wrap, bg=C["bg"])
        scanrow.pack(fill="x")
        tk.Label(scanrow, text="Scan CAC:", font=FSML, fg=C["text"],
                 bg=C["bg"]).pack(side="left")
        self.entry = tk.Entry(scanrow, font=FMONO10, bg=C["surface2"],
                              fg=C["text"], insertbackground=C["text"],
                              relief="flat")
        self.entry.pack(side="left", fill="x", expand=True, padx=(8, 0), ipady=3)
        self.entry.bind("<Return>", self._on_scan)
        self.listbox = tk.Listbox(wrap, height=7, font=FMONO10, bg=C["surface"],
                                  fg=C["text"], selectbackground=C["accent"],
                                  highlightthickness=0, relief="flat",
                                  activestyle="none")
        self.listbox.pack(fill="both", expand=True, pady=(10, 4))
        self.count_lbl = tk.Label(wrap, text="", font=FSML, fg=C["dim"],
                                  bg=C["bg"])
        self.count_lbl.pack(anchor="w")
        btns = tk.Frame(wrap, bg=C["bg"])
        btns.pack(fill="x", pady=(12, 0))
        self.complete_btn = tk.Button(btns, text="Complete Check-Out",
                                      command=self._complete, font=FBOLD,
                                      bg=C["green"], fg="#06210B", relief="flat",
                                      padx=14, pady=6, activebackground=C["green"],
                                      state="disabled", cursor="hand2")
        self.complete_btn.pack(side="left")
        tk.Button(btns, text="Solo Departure (Authorizing Official Override)",
                  command=self._solo_override, font=FSML, bg=C["gold"],
                  fg="#241A00", relief="flat", padx=12, pady=6,
                  activebackground=C["gold"], cursor="hand2").pack(side="left",
                                                                   padx=8)
        tk.Button(btns, text="Cancel", command=self._cancel, font=FSML,
                  bg=C["surface2"], fg=C["text"], relief="flat", padx=12,
                  pady=6, cursor="hand2").pack(side="right")
        if seed_person:
            self._add_member(seed_person)
        self.after(120, lambda: self.entry.focus_set())
        self.bind("<Escape>", lambda e: self._cancel())

    def _add_member(self, person):
        key = person.get("fascn_key", "")
        for m in self.members:
            if key and m.get("fascn_key", "") == key:
                self.app.bell()
                return False
        self.members.append(person)
        nm = (str(person.get("first_name", "")) + " " +
              str(person.get("last_name", ""))).strip() or "(unnamed)"
        unit = person.get("unit", "")
        tag = ("  -  " + unit) if unit else ""
        risk = int(person.get("liberty_risk", 0) or 0)
        suffix = "   [LIBERTY RISK]" if risk else ""
        self.listbox.insert("end", "  " + nm + tag + suffix)
        if risk:
            self.listbox.itemconfig(self.listbox.size() - 1, foreground=C["gold"])
        self._refresh_state()
        return True

    def _refresh_state(self):
        n = len(self.members)
        self.count_lbl.config(text="Personnel departing together: " + str(n))
        if n >= 2 or self.override_sig:
            self.complete_btn.config(state="normal")
        else:
            self.complete_btn.config(state="disabled")

    def _on_scan(self, event=None):
        raw = self.entry.get().strip()
        self.entry.delete(0, "end")
        if not raw:
            return
        key, person = self.app._resolve_scan(raw)
        if not person:
            self.count_lbl.config(
                text="Card not on roster - add via the main window first.")
            self.app.bell()
            return
        person = dict(person)
        person["fascn_key"] = key
        self._add_member(person)

    def _solo_override(self):
        def _signed(ok, name, edipi, reason):
            if not ok:
                return
            self.override_sig = {"name": name, "edipi": edipi, "reason": reason}
            self.complete_btn.config(text="Complete Solo Check-Out")
            self._refresh_state()
        AuthorizingOfficialSignModal(self, "Authorizing Official Override - Solo Departure", _signed)

    def _complete(self):
        if len(self.members) < 2 and not self.override_sig:
            return
        cb = self.on_finalize
        members = list(self.members)
        sig = self.override_sig
        self.destroy()
        cb(members, sig)

    def _cancel(self):
        self.destroy()
        try:
            self.app.status_var.set("Group check-out cancelled.")
        except Exception:
            pass


class GroupCheckInModal(tk.Toplevel):
    def __init__(self, parent, person, members_out, callback):
        super().__init__(parent)
        self.callback = callback
        self.title("Group Check-In")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        wrap = tk.Frame(self, bg=C["bg"], padx=18, pady=16)
        wrap.pack(fill="both", expand=True)
        nm = (str(person.get("first_name", "")) + " " +
              str(person.get("last_name", ""))).strip() or "this person"
        tk.Label(wrap, text="GROUP CHECK-IN", font=FTITLE, fg=C["in_fg"],
                 bg=C["bg"]).pack(anchor="w")
        tk.Label(wrap,
                 text=nm + " checked out with a group. " + str(len(members_out)) +
                      " member(s) of that group are still out.",
                 font=FSML, fg=C["dim"], bg=C["bg"], wraplength=410,
                 justify="left").pack(anchor="w", pady=(2, 8))
        box = tk.Listbox(wrap, height=min(7, max(3, len(members_out))),
                         font=FMONO10, bg=C["surface"], fg=C["text"],
                         highlightthickness=0, relief="flat", activestyle="none")
        box.pack(fill="both", expand=True, pady=(0, 8))
        for m in members_out:
            mn = (str(m.get("first_name", "")) + " " +
                  str(m.get("last_name", ""))).strip() or "(unnamed)"
            box.insert("end", "  " + mn)
        btns = tk.Frame(wrap, bg=C["bg"])
        btns.pack(fill="x")
        tk.Button(btns, text="Check In " + nm + " Only",
                  command=lambda: self._choose("solo"), font=FBOLD,
                  bg=C["accent"], fg="#06210B", relief="flat", padx=12, pady=6,
                  cursor="hand2").pack(side="left")
        tk.Button(btns, text="Check In Entire Group (" + str(len(members_out)) + ")",
                  command=lambda: self._choose("group"), font=FBOLD,
                  bg=C["green"], fg="#06210B", relief="flat", padx=12, pady=6,
                  cursor="hand2").pack(side="left", padx=8)
        tk.Button(btns, text="Cancel", command=self._cancel, font=FSML,
                  bg=C["surface2"], fg=C["text"], relief="flat", padx=12,
                  pady=6, cursor="hand2").pack(side="right")
        self.bind("<Escape>", lambda e: self._cancel())

    def _choose(self, scope):
        cb = self.callback
        self.destroy()
        cb(scope)

    def _cancel(self):
        cb = self.callback
        self.destroy()
        cb(None)


# =============================================================================
# MAIN APPLICATION
# =============================================================================

class CACApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CLDJ CAC Access Control System")
        self.configure(bg=C["bg"])
        self.minsize(1060, 700)

        self.db = RosterDB()
        self._pending_fascn  = None
        self._pending_parsed = None
        self._scan_timer     = None
        self._flash_timer    = None
        self._flash_bar      = None   # set during _build_scan_panel
        self._flash_panel    = None

        self._build_ui()
        self._refresh_log()
        self._refresh_leave_log()
        self._update_roster_count()
        self.after(2000, self._schedule_backup)
        self.bind("<Key>", self._global_key)
        self.scan_entry.focus_set()
        self._tick()

    # ---- UI -----------------------------------------------------------------

    def _build_ui(self):
        tk.Frame(self, bg=C["gold"], height=3).pack(fill="x")

        hdr = tk.Frame(self, bg=C["surface"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="◈  CLDJ CAC Access Control System",
                 font=FTITLE, fg=C["gold"], bg=C["surface"]
                 ).pack(side="left", padx=20)
        tk.Button(hdr, text="Manage Roster", font=FSML,
                  bg=C["surface2"], fg=C["dim"],
                  activebackground=C["border"], relief="flat",
                  padx=12, pady=5, cursor="hand2",
                  command=self._open_roster
                  ).pack(side="right", padx=16)
        self.roster_lbl = tk.Label(hdr, text="Roster: 0 records",
                                    font=FSML, fg=C["dim"], bg=C["surface"])
        self.roster_lbl.pack(side="right", padx=4)
        self.clock_lbl = tk.Label(hdr, text="", font=("Courier New",12),
                                   fg=C["dim"], bg=C["surface"])
        self.clock_lbl.pack(side="right", padx=20)

        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill="both", expand=True, padx=14, pady=10)

        left = tk.Frame(body, bg=C["bg"], width=300)
        left.pack(side="left", fill="y", padx=(0,12))
        left.pack_propagate(False)
        self._build_scan_panel(left)
        self._build_card_panel(left)

        self._build_log_panel(body)

        sb = tk.Frame(self, bg=C["surface"], pady=5)
        sb.pack(fill="x", side="bottom")
        self.status_var = tk.StringVar(
            value="Ready — scan a CAC or type a FASC-N and press Enter.")
        tk.Label(sb, textvariable=self.status_var, font=FSML,
                 fg=C["dim"], bg=C["surface"], anchor="w"
                 ).pack(side="left", padx=12)
        self.count_var = tk.StringVar(value="0 records")
        tk.Label(sb, textvariable=self.count_var, font=FSML,
                 fg=C["dim"], bg=C["surface"]
                 ).pack(side="right", padx=12)

    def _build_scan_panel(self, parent):
        f = tk.Frame(parent, bg=C["surface"],
                     highlightbackground=C["border"], highlightthickness=1)
        f.pack(fill="x", pady=(0,10))

        tk.Label(f, text="SCAN / ENTER FASC-N", font=("Arial",8,"bold"),
                 fg=C["dim"], bg=C["surface"]
                 ).pack(anchor="w", padx=12, pady=(10,4))

        self.scan_var = tk.StringVar()
        self.scan_entry = tk.Entry(
            f, textvariable=self.scan_var, font=FID,
            bg=C["surface2"], fg=C["bright"],
            insertbackground=C["bright"], relief="flat",
            highlightbackground=C["accent"],
            highlightcolor=C["accent_light"],
            highlightthickness=2
        )
        self.scan_entry.pack(fill="x", padx=12, pady=(0,2))
        self.scan_entry.bind("<Return>", self._on_return)
        self.scan_entry.bind("<KeyRelease>", self._on_key_release)

        # Flash bar — 4px animated strip below the scan field
        self._flash_bar = tk.Frame(f, bg=C["border"], height=4)
        self._flash_bar.pack(fill="x", padx=12, pady=(0,2))
        # Keep a ref to the scan panel frame for background flash
        self._flash_panel = f

        tk.Label(f, text="Scan badge or type FASC-N — press Enter",
                 font=("Arial",8), fg=C["dim"], bg=C["surface"]
                 ).pack(anchor="w", padx=12, pady=(0,6))

        btn = tk.Frame(f, bg=C["surface"])
        btn.pack(fill="x", padx=12, pady=(0,12))
        tk.Button(btn, text="CHECK IN", font=FBOLD,
                  bg=C["green"], fg="white", activebackground="#00A368",
                  relief="flat", pady=8, cursor="hand2",
                  command=self._manual_in
                  ).pack(side="left", fill="x", expand=True, padx=(0,6))
        tk.Button(btn, text="CHECK OUT", font=FBOLD,
                  bg=C["red"], fg="white", activebackground="#C02020",
                  relief="flat", pady=8, cursor="hand2",
                  command=self._manual_out
                  ).pack(side="left", fill="x", expand=True)

        self.auto_var = tk.BooleanVar(value=True)
        tk.Checkbutton(f, text="Auto-detect IN/OUT direction",
                       variable=self.auto_var, font=("Arial",8),
                       fg=C["dim"], bg=C["surface"],
                       selectcolor=C["surface2"],
                       activebackground=C["surface"], relief="flat"
                       ).pack(anchor="w", padx=12, pady=(0,10))

    def _build_card_panel(self, parent):
        f = tk.Frame(parent, bg=C["surface"],
                     highlightbackground=C["border"], highlightthickness=1)
        f.pack(fill="both", expand=True)

        tk.Label(f, text="LAST SCANNED CARD", font=("Arial",8,"bold"),
                 fg=C["dim"], bg=C["surface"]
                 ).pack(anchor="w", padx=12, pady=(10,6))

        self._cf = {}
        rows = [
            ("FASC-N",         "fascn_key"),
            ("DoD ID (EDIPI)", "edipi"),
            ("Last Name",      "last_name"),
            ("First Name",     "first_name"),
            ("Middle Name",    "middle_name"),
            ("Date of Birth",  "dob"),
            ("Place of Birth", "place_of_birth"),
            ("Gender",         "gender"),
            ("Branch",         "branch"),
            ("Rank",           "rank"),
            ("Affiliation",    "affiliation"),
            ("Unit / Dept",    "unit"),
            ("Source",         "added_by"),
        ]
        grid = tk.Frame(f, bg=C["surface"])
        grid.pack(fill="x", padx=12)
        for i, (lbl, key) in enumerate(rows):
            tk.Label(grid, text=lbl+":", font=("Arial",8), fg=C["dim"],
                     bg=C["surface"], anchor="w", width=14
                     ).grid(row=i, column=0, sticky="w", pady=1)
            var = tk.StringVar(value="—")
            self._cf[key] = var
            tk.Label(grid, textvariable=var, font=FMONO10,
                     fg=C["text"], bg=C["surface"], anchor="w"
                     ).grid(row=i, column=1, sticky="w")

        dr = tk.Frame(f, bg=C["surface"])
        dr.pack(fill="x", padx=12, pady=(8,10))
        tk.Label(dr, text="Status:", font=FSML, fg=C["dim"],
                 bg=C["surface"]).pack(side="left")
        self.badge = tk.Label(dr, text="  —  ", font=FBOLD,
                              fg=C["dim"], bg=C["surface2"], padx=8, pady=2)
        self.badge.pack(side="left", padx=8)

    def _build_log_panel(self, parent):
        right = tk.Frame(parent, bg=C["bg"])
        right.pack(side="left", fill="both", expand=True)

        # ── Notebook: Access Log | Liberty Risk Log ───────────────────────────
        log_style = ttk.Style(self)
        log_style.configure("LogNB.TNotebook",
                             background=C["bg"], borderwidth=0)
        log_style.configure("LogNB.TNotebook.Tab",
                             background=C["surface2"],
                             foreground=C["dim"],
                             padding=[14, 5],
                             font=("Arial", 9, "bold"))
        log_style.map("LogNB.TNotebook.Tab",
                      background=[("selected", C["accent"])],
                      foreground=[("selected", "white")])

        log_nb = ttk.Notebook(right, style="LogNB.TNotebook")
        log_nb.pack(fill="both", expand=True)

        access_tab = tk.Frame(log_nb, bg=C["bg"])
        lr_tab     = tk.Frame(log_nb, bg=C["bg"])
        leave_tab  = tk.Frame(log_nb, bg=C["bg"])
        log_nb.add(access_tab, text="  ACCESS LOG  ")
        log_nb.add(lr_tab,     text="  ⚠ LIBERTY RISK LOG  ")
        log_nb.add(leave_tab,  text="  🏖 LEAVE STATUS  ")

        self._build_access_log_tab(access_tab)
        self._build_main_lr_log_tab(lr_tab)
        self._build_leave_tab(leave_tab)

    def _build_access_log_tab(self, parent):
        """Standard access log — check-in/out events."""
        hdr = tk.Frame(parent, bg=C["bg"])
        hdr.pack(fill="x", pady=(6, 4), padx=4)
        tk.Label(hdr, text="ACCESS LOG", font=("Arial", 10, "bold"),
                 fg=C["text"], bg=C["bg"]).pack(side="left")
        for lbl, cmd in [("Export XLSX", self._export_xlsx),
                          ("Export CSV",  self._export_csv)]:
            tk.Button(hdr, text=lbl, font=FSML,
                      bg=C["accent"], fg="white",
                      activebackground=C["accent_light"],
                      relief="flat", padx=10, pady=4, cursor="hand2",
                      command=cmd
                      ).pack(side="right", padx=4)
        tk.Button(hdr, text="Clear", font=FSML,
                  bg=C["surface2"], fg=C["dim"],
                  activebackground=C["border"], relief="flat",
                  padx=10, pady=4, cursor="hand2",
                  command=self._clear_log
                  ).pack(side="right", padx=4)

        fb = tk.Frame(parent, bg=C["bg"])
        fb.pack(fill="x", pady=(0, 6), padx=4)
        tk.Label(fb, text="Filter:", font=FSML, fg=C["dim"],
                 bg=C["bg"]).pack(side="left")
        self.filter_var = tk.StringVar()
        tk.Entry(fb, textvariable=self.filter_var, font=FSML,
                 bg=C["surface2"], fg=C["text"], relief="flat",
                 insertbackground=C["text"]
                 ).pack(side="left", padx=6, fill="x", expand=True)
        self.filter_var.trace_add("write", lambda *_: self._refresh_log())
        self.dir_filter = tk.StringVar(value="ALL")
        for val, lbl in [("ALL", "All"), ("IN", "IN"), ("OUT", "OUT")]:
            tk.Radiobutton(fb, text=lbl, variable=self.dir_filter, value=val,
                           font=FSML, fg=C["dim"], bg=C["bg"],
                           selectcolor=C["surface"],
                           activebackground=C["bg"],
                           command=self._refresh_log
                           ).pack(side="left", padx=3)

        tf = tk.Frame(parent, bg=C["bg"])
        tf.pack(fill="both", expand=True, padx=4)

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Log.Treeview",
                        background=C["surface"], foreground=C["text"],
                        fieldbackground=C["surface"], rowheight=26,
                        font=FMONO10)
        style.configure("Log.Treeview.Heading",
                        background=C["surface2"], foreground=C["gold"],
                        font=("Arial", 9, "bold"), relief="flat")
        style.map("Log.Treeview",
                  background=[("selected", C["accent"])],
                  foreground=[("selected", "white")])

        cols = ("Timestamp", "Dir", "FASC-N", "EDIPI",
                "Last Name", "First Name", "Branch", "Rank", "Destination")
        self.tree = ttk.Treeview(tf, columns=cols, show="headings",
                                  style="Log.Treeview", selectmode="browse")
        widths = [155, 55, 175, 100, 110, 100, 70, 60, 160]
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col,
                              command=lambda c=col: self._sort(c))
            self.tree.column(col, width=w, minwidth=50)
        self.tree.tag_configure("IN",
                                background=C["in_bg"], foreground=C["in_fg"])
        self.tree.tag_configure("OUT",
                                background=C["out_bg"], foreground=C["out_fg"])

        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)
        self._sort_col, self._sort_rev = None, False

    def _build_main_lr_log_tab(self, parent):
        """Liberty Risk acknowledgement log — mirrored from RosterWindow."""
        hdr = tk.Frame(parent, bg=C["bg"])
        hdr.pack(fill="x", pady=(6, 4), padx=4)
        tk.Label(hdr, text="LIBERTY RISK LOG", font=("Arial", 10, "bold"),
                 fg="#FF6644", bg=C["bg"]).pack(side="left")
        tk.Button(hdr, text="Export CSV", font=FSML,
                  bg=C["accent"], fg="white",
                  activebackground=C["accent_light"],
                  relief="flat", padx=10, pady=4, cursor="hand2",
                  command=self._export_lr_log_main
                  ).pack(side="right", padx=4)

        fb = tk.Frame(parent, bg=C["bg"])
        fb.pack(fill="x", pady=(0, 6), padx=4)
        tk.Label(fb, text="Filter:", font=FSML, fg=C["dim"],
                 bg=C["bg"]).pack(side="left")
        self._lr_filter_var = tk.StringVar()
        tk.Entry(fb, textvariable=self._lr_filter_var, font=FSML,
                 bg=C["surface2"], fg=C["text"], relief="flat",
                 insertbackground=C["text"]
                 ).pack(side="left", padx=6, fill="x", expand=True)
        self._lr_filter_var.trace_add("write",
                                       lambda *_: self._refresh_lr_log())

        tf = tk.Frame(parent, bg=C["bg"])
        tf.pack(fill="both", expand=True, padx=4)

        style = ttk.Style(self)
        style.configure("LRMain.Treeview",
                        background=C["surface"], foreground=C["text"],
                        fieldbackground=C["surface"], rowheight=26,
                        font=FMONO10)
        style.configure("LRMain.Treeview.Heading",
                        background="#4A1000", foreground="#FF8844",
                        font=("Arial", 9, "bold"), relief="flat")
        style.map("LRMain.Treeview",
                  background=[("selected", "#CC0000")],
                  foreground=[("selected", "white")])

        lr_cols = ("Timestamp", "Last Name", "First Name", "EDIPI",
                   "Direction", "Destination",
                   "Acknowledger Name", "Acknowledger EDIPI", "Reason")
        self.lr_tree_main = ttk.Treeview(tf, columns=lr_cols,
                                          show="headings",
                                          style="LRMain.Treeview",
                                          selectmode="browse")
        lr_widths = [150, 110, 110, 100, 60, 130, 160, 120, 180]
        for col, w in zip(lr_cols, lr_widths):
            self.lr_tree_main.heading(col, text=col)
            self.lr_tree_main.column(col, width=w, minwidth=50)

        vsb = ttk.Scrollbar(tf, orient="vertical",
                             command=self.lr_tree_main.yview)
        self.lr_tree_main.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.lr_tree_main.pack(fill="both", expand=True)

    def _build_leave_tab(self, parent):
        """LEAVE STATUS tab — active and returned leave records."""
        hdr = tk.Frame(parent, bg=C["bg"])
        hdr.pack(fill="x", pady=(6, 4), padx=4)
        tk.Label(hdr, text="LEAVE STATUS", font=("Arial", 10, "bold"),
                 fg=C["gold"], bg=C["bg"]).pack(side="left")
        tk.Button(hdr, text="Export CSV", font=FSML,
                  bg=C["accent"], fg="white",
                  activebackground=C["accent_light"],
                  relief="flat", padx=10, pady=4, cursor="hand2",
                  command=self._export_leave_log
                  ).pack(side="right", padx=4)

        fb = tk.Frame(parent, bg=C["bg"])
        fb.pack(fill="x", pady=(0, 6), padx=4)
        tk.Label(fb, text="Filter:", font=FSML, fg=C["dim"],
                 bg=C["bg"]).pack(side="left")
        self._leave_filter_var = tk.StringVar()
        tk.Entry(fb, textvariable=self._leave_filter_var, font=FSML,
                 bg=C["surface2"], fg=C["text"], relief="flat",
                 insertbackground=C["text"]
                 ).pack(side="left", padx=6, fill="x", expand=True)
        self._leave_filter_var.trace_add("write",
                                         lambda *_: self._refresh_leave_log())
        self._leave_status_filter = tk.StringVar(value="ALL")
        for val, lbl in [("ALL", "All"), ("ACTIVE", "On Leave"),
                         ("RETURNED", "Returned")]:
            tk.Radiobutton(fb, text=lbl, variable=self._leave_status_filter,
                           value=val, font=FSML, fg=C["dim"], bg=C["bg"],
                           selectcolor=C["surface"], activebackground=C["bg"],
                           command=self._refresh_leave_log
                           ).pack(side="left", padx=3)

        tf = tk.Frame(parent, bg=C["bg"])
        tf.pack(fill="both", expand=True, padx=4)

        style = ttk.Style(self)
        style.configure("Leave.Treeview",
                        background=C["surface"], foreground=C["text"],
                        fieldbackground=C["surface"], rowheight=26,
                        font=FMONO10)
        style.configure("Leave.Treeview.Heading",
                        background=C["surface2"], foreground=C["gold"],
                        font=("Arial", 9, "bold"), relief="flat")
        style.map("Leave.Treeview",
                  background=[("selected", C["accent"])],
                  foreground=[("selected", "white")])

        cols = ("Status", "Type", "Last Name", "First Name", "EDIPI",
                "Start Date", "End Date", "Departed", "Returned")
        self.leave_tree = ttk.Treeview(tf, columns=cols, show="headings",
                                       style="Leave.Treeview", selectmode="browse")
        widths = [70, 120, 110, 110, 100, 95, 95, 150, 150]
        for col, w in zip(cols, widths):
            self.leave_tree.heading(col, text=col)
            self.leave_tree.column(col, width=w, minwidth=50)
        self.leave_tree.tag_configure("ACTIVE",
                                      background="#2A2000", foreground=C["gold"])
        self.leave_tree.tag_configure("RETURNED",
                                      background=C["surface"], foreground=C["dim"])

        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.leave_tree.yview)
        self.leave_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.leave_tree.pack(fill="both", expand=True)

    def _refresh_leave_log(self):
        txt = self._leave_filter_var.get() if hasattr(self, "_leave_filter_var") else ""
        status = (self._leave_status_filter.get()
                  if hasattr(self, "_leave_status_filter") else "ALL")
        rows = self.db.get_leave_records(status_filter=status, text_filter=txt)
        for r in self.leave_tree.get_children():
            self.leave_tree.delete(r)
        for row in rows:
            dep = (row.get("departure_ts", "") or "").replace("T", " ")
            ret = (row.get("return_ts", "") or "").replace("T", " ") or "—"
            self.leave_tree.insert("", "end", values=(
                row.get("status", ""),
                row.get("leave_type", ""),
                row.get("last_name", ""),
                row.get("first_name", ""),
                row.get("edipi", ""),
                row.get("start_date", ""),
                row.get("end_date", ""),
                dep,
                ret,
            ), tags=(row.get("status", ""),))

    def _export_leave_log(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("All", "*.*")],
            initialfile=f"leave_records_{datetime.date.today()}.csv"
        )
        if not path:
            return
        rows = self.db.get_leave_records()
        cols = ["status", "leave_type", "last_name", "first_name", "edipi",
                "start_date", "end_date", "departure_ts", "return_ts", "fascn_key"]
        with open(path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)
        self.status_var.set(f"Leave records exported: {path}")

    def _refresh_lr_log(self):
        txt = self._lr_filter_var.get() if hasattr(self, "_lr_filter_var") else ""
        rows = self.db.get_liberty_risk_log(text_filter=txt)
        for r in self.lr_tree_main.get_children():
            self.lr_tree_main.delete(r)
        for row in rows:
            self.lr_tree_main.insert("", "end", values=(
                row["timestamp"],
                row["last_name"],
                row["first_name"],
                row["edipi"],
                row["direction"],
                row["destination"],
                row["acknowledger_name"],
                row["acknowledger_edipi"],
                row["liberty_risk_reason"],
            ))
        total = self.db.liberty_risk_log_count()
        self.count_var.set(
            f"Access: {self.db.total_log_count()} records  |  "
            f"Liberty Risk: {total} acknowledgements"
        )

    def _export_lr_log_main(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("All", "*.*")],
            initialfile=f"liberty_risk_log_{datetime.date.today()}.csv"
        )
        if not path:
            return
        rows = self.db.get_liberty_risk_log()
        cols = ["timestamp", "last_name", "first_name", "edipi",
                "direction", "destination", "acknowledger_name",
                "acknowledger_edipi", "liberty_risk_reason"]
        with open(path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)
        self.status_var.set(f"Liberty Risk log exported: {path}")

    # ---- Scan handling ------------------------------------------------------

    def _global_key(self, event):
        if self.focus_get() is self.scan_entry:
            return
        if event.char and event.char.isprintable():
            self.scan_var.set(self.scan_var.get() + event.char)
            self.scan_entry.focus_set()
            self.scan_entry.icursor(tk.END)
        elif event.keysym in ("Return","KP_Enter"):
            self._on_return(event)

    def _on_key_release(self, event):
        # Primary path: HID wedge reader sends CR after burst — _on_return fires instantly.
        # Idle timer is a fallback for readers that don't send CR, or manual typing.
        # Trigger threshold = 18 chars (confirmed FASC-N length for this reader).
        # Timer = 150ms (tighter than default 220ms; NXP burst completes in <100ms).
        if self._scan_timer:
            self.after_cancel(self._scan_timer)
        raw = self.scan_var.get()
        if len(raw) >= 18:
            self._scan_timer = self.after(150, self._auto_scan)

    def _auto_scan(self):
        raw = self.scan_var.get().strip()
        if raw and self.auto_var.get():
            self._process_scan(raw)
            self.scan_var.set("")

    def _on_return(self, _=None):
        # CR from reader — cancel any pending idle timer and process immediately.
        if self._scan_timer:
            self.after_cancel(self._scan_timer)
            self._scan_timer = None
        raw = self.scan_var.get().strip()
        if raw:
            self._process_scan(raw)
        self.scan_var.set("")

    def _process_scan(self, raw: str):
        parsed = FASCNParser.parse(raw)
        key    = parsed["fascn_key"]
        person = self.db.lookup(key)
        front  = parsed.get("source") == "front_PDF417"

        # EDIPI bridge: if FASC-N not found, check if the parsed EDIPI
        # matches an EDIPI-keyed import record (key = 'EDIPI:<edipi>')
        if not person and parsed.get("edipi"):
            person = self.db.lookup_by_edipi(parsed["edipi"])
            if person:
                self.db.link_fascn_to_edipi(key, parsed["edipi"])

        if person:
            direction = (self._infer_dir(key)
                         if self.auto_var.get() else "IN")
            self._record(key, person, direction,
                         suffix=" [front barcode]" if front else "")
        else:
            self._pending_fascn  = key
            self._pending_parsed = parsed
            side = "front" if front else "back"
            self.status_var.set(
                f"Unknown card ({side} scan) — FASC-N: {key[:18]}  Enter details.")
            UnknownCardModal(self, key, parsed, self._on_manual_save, self.db)

    def _on_manual_save(self, data: dict, direction: str, link_fascn: str = None):
        """
        Callback from UnknownCardModal.

        Two paths:
          link_fascn is set  — operator confirmed an EDIPI lookup match.
                               Link the scanned FASC-N to the existing record,
                               then log the event under the new key.
          link_fascn is None — operator filled in manual details.
                               Create a new record keyed by the scanned FASC-N.
        """
        key = self._pending_fascn
        if not key:
            return

        if link_fascn:
            # Link the scanned FASC-N to the existing DB record
            edipi = data.get("edipi", "")
            self.db.link_fascn_to_edipi(link_fascn, edipi)
            # After linking the key has changed — look up by the new FASC-N
            person = self.db.lookup(link_fascn)
            if not person:
                # Fallback: use the data dict we got back from the modal
                person = data
            record_key = link_fascn
            self.status_var.set(
                f"FASC-N linked to {data.get('first_name','')} "
                f"{data.get('last_name','')} (EDIPI: {edipi})")
        else:
            # New record — upsert with the scanned FASC-N as the key
            self.db.upsert_person(key, data, added_by="manual")
            person = self.db.lookup(key)
            record_key = key

        self._record(record_key, person, direction)
        self._update_roster_count()
        self._pending_fascn = self._pending_parsed = None

    def _manual_in(self):
        raw = self.scan_var.get().strip()
        if not raw:
            self.status_var.set("Scan or enter a FASC-N first.")
            return
        parsed = FASCNParser.parse(raw)
        key    = parsed["fascn_key"]
        person = self.db.lookup(key)
        if person:
            self._record(key, person, "IN")
        else:
            self._pending_fascn  = key
            self._pending_parsed = parsed
            UnknownCardModal(self, key, parsed, self._on_manual_save, self.db)
        self.scan_var.set("")

    def _manual_out(self):
        raw = self.scan_var.get().strip()
        if not raw:
            self.status_var.set("Scan or enter a FASC-N first.")
            return
        parsed = FASCNParser.parse(raw)
        key    = parsed["fascn_key"]
        person = self.db.lookup(key)
        if person:
            self._record(key, person, "OUT")
        else:
            self._pending_fascn  = key
            self._pending_parsed = parsed
            UnknownCardModal(self, key, parsed, self._on_manual_save, self.db)
        self.scan_var.set("")

    def _infer_dir(self, key: str) -> str:
        last = self.db.last_direction(key)
        return "IN" if last == "OUT" else "OUT"

    def _record(self, key, person, direction, suffix=""):
        missing = self._missing_critical_fields(person)
        if missing:
            def _continue():
                updated = self.db.lookup(key) or person
                self._record_dispatch(key, updated, direction, suffix)
            RecordUpdateModal(self, person, missing, self.db, _continue)
            return
        self._record_dispatch(key, person, direction, suffix)

    def _missing_critical_fields(self, person):
        crit = [("DoD ID (EDIPI)", "edipi"), ("Last Name", "last_name"),
                ("First Name", "first_name"), ("Unit / Department", "unit")]
        return [label for label, k in crit
                if not str(person.get(k, "") or "").strip()]

    def _resolve_scan(self, raw):
        parsed = FASCNParser.parse(raw)
        key = parsed["fascn_key"]
        person = self.db.lookup(key)
        if not person and parsed.get("edipi"):
            person = self.db.lookup_by_edipi(parsed["edipi"])
            if person:
                self.db.link_fascn_to_edipi(key, parsed["edipi"])
        return key, person

    def _record_dispatch(self, key, person, direction, suffix=""):
        if direction == "OUT":
            self._group_checkout(key, person, suffix)
        else:
            self._checkin_with_group(key, person, suffix)

    def _group_checkout(self, key, person, suffix=""):
        seed = dict(person)
        seed["fascn_key"] = key

        def _finalize(members, override_sig):
            if not members:
                self.status_var.set("Group check-out cancelled.")
                return
            group_id = "G" + datetime.datetime.now().strftime("%Y%m%d%H%M%S")
            risk_members = [m for m in members
                            if int(m.get("liberty_risk", 0) or 0)]

            def _proceed(ack_name, ack_edipi):
                if risk_members and ack_name:
                    for m in risk_members:
                        self.db.log_liberty_risk_acknowledgement(
                            m.get("fascn_key", ""), m, "OUT", "",
                            ack_name, ack_edipi)

                def _on_dest(dest):
                    leave_type = (LEAVE_RR if dest == "R&R Leave"
                                  else LEAVE_96 if dest == "96-Hour Liberty"
                                  else None)
                    if leave_type:
                        def _on_dates(start_iso, end_iso):
                            if not start_iso:
                                self.status_var.set("Group leave cancelled.")
                                return
                            full = (leave_type + " (" + start_iso + " to " +
                                    end_iso + ")")
                            for m in members:
                                mk = m.get("fascn_key", "")
                                self.db.create_leave_record(mk, m, leave_type,
                                                            start_iso, end_iso)
                                self.db.log_event("OUT", mk, m, destination=full,
                                                  group_id=group_id)
                            self._after_group_out(members, group_id, full)
                        LeaveDateRangeModal(self, seed, leave_type, _on_dates)
                        return
                    for m in members:
                        mk = m.get("fascn_key", "")
                        self.db.log_event("OUT", mk, m, destination=dest,
                                          group_id=group_id)
                    self._after_group_out(members, group_id, dest)
                DestinationModal(self, seed, _on_dest)

            if override_sig:
                _proceed(override_sig.get("name", ""),
                         override_sig.get("edipi", ""))
            elif risk_members:
                def _signed(ok, name, edipi, reason):
                    if not ok:
                        self.status_var.set("Group check-out cancelled "
                                            "(liberty risk not acknowledged).")
                        return
                    _proceed(name, edipi)
                AuthorizingOfficialSignModal(
                    self,
                    "Liberty Risk in group - authorizing official acknowledgement required",
                    _signed)
            else:
                _proceed("", "")

        GroupCheckOutModal(self, seed, self.db, _finalize)

    def _after_group_out(self, members, group_id, dest_label):
        self._refresh_log()
        self._refresh_leave_log()
        self._flash_event("OUT")
        last = members[-1]
        self._show_card(last.get("fascn_key", ""), last, "OUT")
        names = ", ".join(
            ((str(m.get("first_name", "")) + " " +
              str(m.get("last_name", ""))).strip() or "(unnamed)")
            for m in members)
        self.status_var.set("GROUP OUT (" + str(len(members)) + "): " + names +
                            "  |  " + dest_label)

    def _checkin_with_group(self, key, person, suffix=""):
        gid = self.db.last_group_id(key)
        members_out = self.db.group_members_out(gid) if gid else []
        others = [m for m in members_out if m.get("fascn_key", "") != key]
        if gid and others:
            def _choice(scope):
                if scope is None:
                    self.status_var.set("Group check-in cancelled.")
                    return
                if scope == "solo":
                    self._record_proceed(key, person, "IN", suffix)
                else:
                    for m in members_out:
                        self._record_proceed(m.get("fascn_key", ""), m, "IN",
                                             suffix)
            GroupCheckInModal(self, person, members_out, _choice)
        else:
            self._record_proceed(key, person, "IN", suffix)

    def _record_proceed(self, key: str, person: dict, direction: str, suffix: str = ""):
        """
        Log a check-in or check-out event.

        Liberty Risk persons trigger LibertyRiskModal before any logging occurs.
        For OUT events, DestinationModal is shown first, then Liberty Risk check.
        """
        is_liberty_risk = bool(person.get("liberty_risk", 0))

        if direction == "OUT":
            def _on_dest(dest: str):
                leave_type = (LEAVE_RR if dest == "R&R Leave"
                              else LEAVE_96 if dest == "96-Hour Liberty"
                              else None)
                if leave_type:
                    def _on_dates(start_iso, end_iso):
                        if not start_iso:
                            self.status_var.set("Leave entry cancelled.")
                            return
                        full_dest = f"{leave_type} ({start_iso} to {end_iso})"
                        def _commit_leave():
                            self.db.create_leave_record(key, person, leave_type,
                                                        start_iso, end_iso)
                            entry = self.db.log_event(direction, key, person,
                                                      destination=full_dest)
                            self._show_card(key, person, direction)
                            self._refresh_log()
                            self._refresh_leave_log()
                            self._flash_event(direction)
                            name = (f"{person.get('first_name','')} "
                                    f"{person.get('last_name','')}").strip()
                            self.status_var.set(
                                f"{leave_type}: {name}  |  "
                                f"{start_iso} → {end_iso}  |  {entry['time']}")
                        if is_liberty_risk:
                            def _on_lr_leave(acknowledged, sig_name, sig_edipi):
                                if not acknowledged:
                                    self.status_var.set(
                                        "Liberty Risk leave CANCELLED by operator.")
                                    return
                                _commit_leave()
                            LibertyRiskModal(self, person, direction, full_dest,
                                             self.db, _on_lr_leave)
                        else:
                            _commit_leave()
                    LeaveDateRangeModal(self, person, leave_type, _on_dates)
                    return
                if is_liberty_risk:
                    def _on_lr(acknowledged: bool,
                               sig_name: str, sig_edipi: str):
                        if not acknowledged:
                            self.status_var.set(
                                "Liberty Risk check-out CANCELLED by operator.")
                            return
                        entry = self.db.log_event(
                            direction, key, person, destination=dest)
                        self._show_card(key, person, direction)
                        self._refresh_log()
                        self._refresh_lr_log()
                        self._flash_event(direction)
                        name = (f"{person.get('first_name','')} "
                                f"{person.get('last_name','')}").strip()
                        self.status_var.set(
                            f"CHECK OUT [LIBERTY RISK]: {name}  |  "
                            f"{entry['time']}  →  {dest}  |  "
                            f"Ack: {sig_name}"
                        )
                    LibertyRiskModal(self, person, direction,
                                     dest, self.db, _on_lr)
                else:
                    entry = self.db.log_event(
                        direction, key, person, destination=dest)
                    self._show_card(key, person, direction)
                    self._refresh_log()
                    self._flash_event(direction)
                    name = (f"{person.get('first_name','')} "
                            f"{person.get('last_name','')}").strip()
                    dest_sfx = f"  →  {dest}" if dest else ""
                    self.status_var.set(
                        f"CHECK OUT: {name}  |  "
                        f"{entry['time']}{dest_sfx}{suffix}"
                    )
            DestinationModal(self, person, _on_dest)

        else:  # IN
            if is_liberty_risk:
                def _on_lr_in(acknowledged: bool,
                               sig_name: str, sig_edipi: str):
                    if not acknowledged:
                        self.status_var.set(
                            "Liberty Risk check-in CANCELLED by operator.")
                        return
                    entry = self.db.log_event(direction, key, person)
                    self._show_card(key, person, direction)
                    self._refresh_log()
                    self._refresh_lr_log()
                    self._flash_event(direction)
                    name = (f"{person.get('first_name','')} "
                            f"{person.get('last_name','')}").strip()
                    self.status_var.set(
                        f"CHECK IN [LIBERTY RISK]: {name}  |  "
                        f"{entry['time']}  |  Ack: {sig_name}"
                    )
                LibertyRiskModal(self, person, direction,
                                 "", self.db, _on_lr_in)
            else:
                # Check for an ACTIVE leave record — confirm return if found
                active_leave = self.db.get_active_leave(key)
                if active_leave:
                    def _on_confirm(confirmed):
                        if not confirmed:
                            self.status_var.set(
                                "Return from leave CANCELLED — still on leave.")
                            return
                        self.db.mark_leave_returned(active_leave["id"])
                        entry = self.db.log_event(direction, key, person)
                        self._show_card(key, person, direction)
                        self._refresh_log()
                        self._refresh_leave_log()
                        self._flash_event(direction)
                        name = (f"{person.get('first_name','')} "
                                f"{person.get('last_name','')}").strip()
                        self.status_var.set(
                            f"RETURNED FROM LEAVE: {name}  |  "
                            f"{active_leave['leave_type']}  |  {entry['time']}")
                    LeaveReturnModal(self, person, active_leave, _on_confirm)
                    return
                entry = self.db.log_event(direction, key, person)
                self._show_card(key, person, direction)
                self._refresh_log()
                self._flash_event(direction)
                name = (f"{person.get('first_name','')} "
                        f"{person.get('last_name','')}").strip()
                self.status_var.set(
                    f"CHECK IN: {name}  |  FASC-N: {key[:20]}  |  "
                    f"{entry['time']}{suffix}"
                )

    def _flash_event(self, direction: str):
        """
        Three-layer visual flash on each scan event:
          1. Flash bar (4px strip under scan field) — pulses bright then fades
          2. Scan entry border — briefly highlights in event colour
          3. Status bar background — quick colour wash
        Uses tkinter .after() chained steps; no threads needed.
        """
        flash_col  = C["green"] if direction == "IN" else C["red"]
        dim_col    = "#007040" if direction == "IN" else "#801010"
        idle_col   = C["border"]

        # Cancel any in-progress flash
        if self._flash_timer:
            self.after_cancel(self._flash_timer)

        # Step sequence: (delay_ms, bar_colour, entry_border, panel_bg)
        steps = [
            (0,   flash_col, flash_col, C["surface2"]),
            (80,  flash_col, flash_col, C["surface"]),
            (160, dim_col,   C["accent"], C["surface"]),
            (280, dim_col,   C["accent"], C["surface"]),
            (420, idle_col,  C["accent"], C["surface"]),
        ]

        def apply_step(idx):
            if idx >= len(steps):
                # Fully restore
                self._flash_bar.config(bg=idle_col)
                self.scan_entry.config(
                    highlightbackground=C["accent"],
                    highlightcolor=C["accent_light"]
                )
                self._flash_timer = None
                return
            delay, bar, border, panel = steps[idx]
            self._flash_bar.config(bg=bar)
            self.scan_entry.config(
                highlightbackground=border,
                highlightcolor=border
            )
            self._flash_panel.config(bg=panel)
            self._flash_timer = self.after(
                steps[idx + 1][0] - delay if idx + 1 < len(steps) else 200,
                lambda: apply_step(idx + 1)
            )

        apply_step(0)

    # ---- Display ------------------------------------------------------------

    def _show_card(self, key: str, person: dict, direction: str):
        display_key = key[:24] + ("..." if len(key) > 24 else "")
        self._cf["fascn_key"].set(display_key)
        for field in ["edipi","last_name","first_name","middle_name",
                      "dob","place_of_birth","gender","branch",
                      "rank","affiliation","unit","added_by"]:
            self._cf[field].set(person.get(field, "—") or "—")
        col = C["green"] if direction == "IN" else C["red"]
        self.badge.config(text=f"  {direction}  ", bg=col, fg="white")

    def _backup_logs(self):
        try:
            bdir = pathlib.Path("log_backups")
            bdir.mkdir(exist_ok=True)
            self._write_backup_readme(bdir)
            stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            jobs = [
                ("access_log", self.db.get_log()),
                ("roster_audit_log", self.db.get_audit_log()),
                ("liberty_risk_log", self.db.get_liberty_risk_log()),
                ("leave_records", self.db.get_leave_records()),
            ]
            written = 0
            for name, rows in jobs:
                path = bdir / (name + "_" + stamp + ".csv")
                cols = []
                for r in rows:
                    for k in r.keys():
                        if k not in cols:
                            cols.append(k)
                with open(path, "w", newline="", encoding="utf-8") as f:
                    w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
                    w.writeheader()
                    for r in rows:
                        w.writerow(dict(r))
                try:
                    os.chmod(path, 0o444)
                except Exception:
                    pass
                written += 1
            self._prune_backups(bdir, keep=80)
            return written
        except Exception as exc:
            try:
                self.status_var.set("Log backup error: " + str(exc))
            except Exception:
                pass
            return 0

    def _write_backup_readme(self, bdir):
        readme = bdir / "SECURITY_README.txt"
        if readme.exists():
            return
        try:
            txt = (
                "AUDIT-LOG BACKUPS - HANDLING (CUI)\n"
                "==================================\n\n"
                "This folder holds automated, timestamped CSV copies of the\n"
                "application audit trail (access log, roster changes, liberty-risk\n"
                "acknowledgements, leave records). Treat as Controlled Unclassified\n"
                "Information (CUI). Files are written READ-ONLY by the application.\n\n"
                "Protection model on this workstation:\n"
                "  1. At-rest encryption is provided by BitLocker full-disk\n"
                "     encryption on this endpoint.\n"
                "  2. Files are set read-only to prevent casual modification.\n"
                "  3. Access should be restricted to authorized accounts. An\n"
                "     administrator can lock this folder down with (run once,\n"
                "     elevated; replace ACCOUNT with the authorized user/group):\n\n"
                "       icacls \"%CD%\\log_backups\" /inheritance:r ^\n"
                "         /grant:r \"ACCOUNT:(OI)(CI)R\" ^\n"
                "         /grant:r \"Administrators:(OI)(CI)F\"\n\n"
                "NOTE: The application is standard-library only and cannot create\n"
                "AES-encrypted (password) archives without a host tool. The\n"
                "read-only attribute is deterrence, NOT encryption. The real\n"
                "confidentiality controls are BitLocker (encryption) plus the NTFS\n"
                "ACL above (access control). Do not rely on the read-only bit alone.\n"
            )
            with open(readme, "w", encoding="utf-8") as f:
                f.write(txt)
            try:
                os.chmod(readme, 0o444)
            except Exception:
                pass
        except Exception:
            pass

    def _prune_backups(self, bdir, keep=80):
        try:
            files = sorted(bdir.glob("*.csv"),
                           key=lambda p: p.stat().st_mtime, reverse=True)
            for old in files[keep:]:
                try:
                    os.chmod(old, 0o666)
                    old.unlink()
                except Exception:
                    pass
        except Exception:
            pass

    def _schedule_backup(self):
        n = self._backup_logs()
        if n:
            try:
                ts = datetime.datetime.now().strftime("%H:%M:%S")
                self.status_var.set("Audit logs backed up (" + str(n) +
                                    " read-only files) at " + ts)
            except Exception:
                pass
        self.after(30 * 60 * 1000, self._schedule_backup)

    def _tick(self):
        now = datetime.datetime.now().strftime("%Y-%m-%d   %H:%M:%S")
        self.clock_lbl.config(text=now)
        self.after(1000, self._tick)

    def _update_roster_count(self):
        n = self.db.roster_count()
        self.roster_lbl.config(text=f"Roster: {n:,} records")

    # ---- Log ----------------------------------------------------------------

    def _refresh_log(self):
        df  = self.dir_filter.get()
        txt = self.filter_var.get()
        rows = self.db.get_log(df, txt)
        for r in self.tree.get_children():
            self.tree.delete(r)
        for row in rows:
            self.tree.insert("", "end", values=(
                row["timestamp"],
                row["direction"],
                row["fascn_key"][:20],
                row["edipi"],
                row["last_name"],
                row["first_name"],
                row["branch"],
                row["rank"],
                row.get("destination", ""),
            ), tags=(row["direction"],))
        total     = self.db.total_log_count()
        lr_total  = self.db.liberty_risk_log_count()
        self.count_var.set(
            f"Access: {len(rows)} shown / {total} total  |  "
            f"Liberty Risk: {lr_total} acknowledgements"
        )

    def _sort(self, col):
        data = [(self.tree.set(c, col), c) for c in self.tree.get_children("")]
        self._sort_rev = (not self._sort_rev
                          if self._sort_col == col else False)
        self._sort_col = col
        data.sort(reverse=self._sort_rev)
        for i, (_, c) in enumerate(data):
            self.tree.move(c, "", i)

    def _clear_log(self):
        if messagebox.askyesno("Clear Log",
                               "Delete ALL log entries?\nThis cannot be undone.",
                               icon="warning"):
            self.db.clear_log()
            self._refresh_log()
            self.status_var.set("Log cleared.")

    # ---- Exports ------------------------------------------------------------

    def _export_csv(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV","*.csv"),("All","*.*")],
            initialfile=f"cac_log_{datetime.date.today()}.csv"
        )
        if path:
            self.db.export_csv(path)
            self.status_var.set(f"CSV saved: {path}")

    def _export_xlsx(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx"),("All","*.*")],
            initialfile=f"cac_log_{datetime.date.today()}.xlsx"
        )
        if path:
            self.db.export_xlsx(path)
            self.status_var.set(f"XLSX saved: {path}")

    def _open_roster(self):
        win = RosterWindow(self, self.db, self._update_roster_count)
        win.after(100, win._load_audit)
        win.after(120, win._load_lr_log)
        win.after(140, win._load_roster_leave)


# =============================================================================
if __name__ == "__main__":
    app = CACApp()
    app.mainloop()
